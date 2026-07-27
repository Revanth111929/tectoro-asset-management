# routes.py - Simplified routes for IT Asset Management
import os, csv, io
from datetime import datetime, date, timedelta
from flask import (Blueprint, render_template, redirect, url_for,
                   request, flash, send_file, jsonify, current_app)
# flask_login session imports removed — app uses JWT token auth
# from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import or_, func
from models import db, User, Asset, ActivityLog
from services.audit_service import AuditService, LifecycleService
from email_service import send_acknowledgment_email
from flask_cors import cross_origin
from werkzeug.utils import secure_filename
# ── RBAC helpers ──────────────────────────────────────────────────────────────
def get_request_user():
    """Extract user from Bearer token."""
    auth = request.headers.get('Authorization', '')
    token = auth.replace('Bearer ', '')
    if token.startswith('user-'):
        parts = token.split('-')
        if len(parts) >= 3:
            try:
                return User.query.get(int(parts[1]))
            except Exception:
                pass
    return None

def require_role(*allowed_roles):
    """Decorator: deny request if user role not in allowed_roles."""
    from functools import wraps
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            user = get_request_user()
            if not user:
                return jsonify({'error': 'Authentication required'}), 401
            if user.role not in allowed_roles:
                return jsonify({'error': 'Access denied: insufficient permissions'}), 403
            return f(*args, **kwargs)
        return wrapped
    return decorator

def require_not_viewer(f):
    """Decorator: deny if viewer role."""
    from functools import wraps
    @wraps(f)
    def wrapped(*args, **kwargs):
        user = get_request_user()
        if user and user.role == 'viewer':
            return jsonify({'error': 'Access denied: View Only users cannot perform this action'}), 403
        return f(*args, **kwargs)
    return wrapped



# Helper: write to activity log
def log_activity(action, module, description, user='admin'):
    """Write an activity log entry — no flask_login dependency."""
    try:
        entry = ActivityLog(user=user, action=action, module=module, description=description)
        db.session.add(entry)
        db.session.commit()
    except Exception:
        db.session.rollback()

def _parse_date(val):
    if val:
        try:
            return datetime.strptime(val, '%Y-%m-%d').date()
        except ValueError:
            pass
    return None

# AUTH BLUEPRINT
auth_bp = Blueprint('auth', __name__)

@auth_bp.route('/')
def landing():
    return jsonify({'message': 'Tectoro Asset Management API'}), 200

@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    # HTML login replaced by React frontend + /api/auth/login JSON endpoint
    return jsonify({'message': 'Use POST /api/auth/login'}), 200

@auth_bp.route('/logout')
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('auth.login'))

# MAIN BLUEPRINT
main_bp = Blueprint('main', __name__)

@main_bp.route('/dashboard')
def dashboard():
    total_assets     = Asset.query.count()
    assigned_assets  = Asset.query.filter_by(status='Assigned').count()
    available_assets = Asset.query.filter_by(status='Available').count()
    maintenance      = Asset.query.filter_by(status='Maintenance').count()

    recent_logs = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).limit(10).all()

    # Assets by category
    cat_data = db.session.query(Asset.category, func.count(Asset.id))\
        .filter(Asset.category != None)\
        .group_by(Asset.category).all()
    cat_labels = [c[0] for c in cat_data]
    cat_counts = [c[1] for c in cat_data]

    # Warranty expiring within 90 days
    today = date.today()
    soon = today + timedelta(days=90)
    expiring = Asset.query.filter(
        Asset.warranty_date != None,
        Asset.warranty_date <= soon,
        Asset.warranty_date >= today
    ).count()

    return render_template('dashboard.html',
        total_assets=total_assets,
        assigned_assets=assigned_assets,
        available_assets=available_assets,
        maintenance=maintenance,
        recent_logs=recent_logs,
        cat_labels=cat_labels,
        cat_counts=cat_counts,
        expiring_soon=expiring
    )

# ASSET BLUEPRINT
asset_bp = Blueprint('asset', __name__, url_prefix='/assets')

@asset_bp.route('/')
def list_assets():
    search   = request.args.get('search', '')
    category = request.args.get('category', '')
    status   = request.args.get('status', '')
    location = request.args.get('location', '')
    page     = request.args.get('page', 1, type=int)

    query = Asset.query

    if search:
        query = query.filter(or_(
            Asset.asset_name.ilike(f'%{search}%'),
            Asset.serial_number.ilike(f'%{search}%'),
            Asset.emp_id.ilike(f'%{search}%'),
            Asset.employee_name.ilike(f'%{search}%')
        ))
    if category:
        query = query.filter_by(category=category)
    if status:
        query = query.filter_by(status=status)
    if location:
        query = query.filter_by(location=location)

    assets = query.order_by(Asset.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    
    categories = db.session.query(Asset.category).distinct().filter(Asset.category != None).all()
    locations = db.session.query(Asset.location).distinct().filter(Asset.location != None).all()

    return render_template('assets/list.html',
        assets=assets,
        categories=[c[0] for c in categories],
        locations=[l[0] for l in locations],
        search=search, sel_category=category, sel_status=status, sel_location=location
    )

@asset_bp.route('/add', methods=['GET', 'POST'])
def add_asset():
    if request.method == 'POST':
        serial = request.form['serial_number'].strip()
        
        asset = Asset(
            emp_id           = request.form.get('emp_id', '').strip(),
            employee_name    = request.form.get('employee_name', '').strip(),
            mobile_number    = request.form.get('mobile_number', '').strip(),
            asset_name       = request.form['asset_name'].strip(),
            category         = request.form.get('category', '').strip(),
            serial_number    = serial,
            model_name       = request.form.get('model_name', '').strip(),
            os               = request.form.get('os', '').strip(),
            version          = request.form.get('version', '').strip(),
            ram              = request.form.get('ram', '').strip(),
            location         = request.form.get('location', '').strip(),
            invoice_number   = request.form.get('invoice_number', '').strip(),
            invoice_date     = _parse_date(request.form.get('invoice_date')),
            warranty_date    = _parse_date(request.form.get('warranty_date')),
            charger_serial   = request.form.get('charger_serial', '').strip(),
            old_user         = request.form.get('old_user', '').strip(),
            date             = _parse_date(request.form.get('date')) or date.today(),
            old_device       = request.form.get('old_device', '').strip(),
            comments         = request.form.get('comments', '').strip(),
            status           = request.form.get('status', 'Available')
        )
        db.session.add(asset)
        db.session.commit()
        log_activity('CREATE', 'Asset', f'Added asset: {asset.asset_name} [{serial}]')
        flash(f'Asset "{asset.asset_name}" added successfully!', 'success')
        return redirect(url_for('asset.list_assets'))

    return render_template('assets/add.html')

@asset_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_asset(id):
    asset = Asset.query.get_or_404(id)

    if request.method == 'POST':
        asset.emp_id         = request.form.get('emp_id', '').strip()
        asset.employee_name  = request.form.get('employee_name', '').strip()
        asset.mobile_number  = request.form.get('mobile_number', '').strip()
        asset.asset_name     = request.form['asset_name'].strip()
        asset.category       = request.form.get('category', '').strip()
        asset.serial_number  = request.form['serial_number'].strip()
        asset.model_name     = request.form.get('model_name', '').strip()
        asset.os             = request.form.get('os', '').strip()
        asset.version        = request.form.get('version', '').strip()
        asset.ram            = request.form.get('ram', '').strip()
        asset.location       = request.form.get('location', '').strip()
        asset.invoice_number = request.form.get('invoice_number', '').strip()
        asset.invoice_date   = _parse_date(request.form.get('invoice_date'))
        asset.warranty_date  = _parse_date(request.form.get('warranty_date'))
        asset.charger_serial = request.form.get('charger_serial', '').strip()
        asset.old_user       = request.form.get('old_user', '').strip()
        asset.date           = _parse_date(request.form.get('date'))
        asset.old_device     = request.form.get('old_device', '').strip()
        asset.comments       = request.form.get('comments', '').strip()
        asset.status         = request.form.get('status', 'Available')
        asset.updated_at     = datetime.utcnow()
        db.session.commit()
        log_activity('UPDATE', 'Asset', f'Updated asset: {asset.asset_name}')
        flash('Asset updated successfully!', 'success')
        return redirect(url_for('asset.list_assets'))

    return render_template('assets/edit.html', asset=asset)

@asset_bp.route('/delete/<int:id>', methods=['POST'])
def delete_asset(id):
    asset = Asset.query.get_or_404(id)
    name  = asset.asset_name
    db.session.delete(asset)
    db.session.commit()
    log_activity('DELETE', 'Asset', f'Deleted asset: {name}')
    flash(f'Asset "{name}" deleted.', 'warning')
    return redirect(url_for('asset.list_assets'))

@asset_bp.route('/view/<int:id>')
def view_asset(id):
    asset = Asset.query.get_or_404(id)
    return render_template('assets/view.html', asset=asset)

# REPORT BLUEPRINT
report_bp = Blueprint('report', __name__, url_prefix='/reports')

@report_bp.route('/')
def reports():
    return render_template('reports/index.html')

@report_bp.route('/export/csv')
def export_csv():
    assets = Asset.query.all()
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(['Sl No','EMP ID','Employee Name','Mobile','Asset Name','Category',
                     'Serial Number','Model','OS','Version','RAM','Location',
                     'Invoice Number','Invoice Date','Warranty Date','Charger Serial',
                     'Old User','Date','Old Device','Comments','Status'])

    for a in assets:
        writer.writerow([
            a.id, a.emp_id or '', a.employee_name or '', a.mobile_number or '',
            a.asset_name, a.category or '', a.serial_number, a.model_name or '',
            a.os or '', a.version or '', a.ram or '', a.location or '',
            a.invoice_number or '', a.invoice_date or '', a.warranty_date or '',
            a.charger_serial or '', a.old_user or '', a.date or '',
            a.old_device or '', a.comments or '', a.status
        ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'assets_{date.today()}.csv'
    )

@report_bp.route('/activity')
def activity_log():
    page = request.args.get('page', 1, type=int)
    logs = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('reports/activity.html', logs=logs)

@report_bp.route('/warranty')
def warranty_alerts():
    today = date.today()
    soon = today + timedelta(days=90)
    expiring = Asset.query.filter(
        Asset.warranty_date != None,
        Asset.warranty_date <= soon,
        Asset.warranty_date >= today
    ).order_by(Asset.warranty_date).all()
    return render_template('reports/warranty.html', assets=expiring)

# API BLUEPRINT (for React frontend)
api_bp = Blueprint('api', __name__, url_prefix='/api')

@api_bp.route('/dashboard/stats')
@cross_origin()
def api_dashboard_stats():
    total     = Asset.query.count()
    assigned  = Asset.query.filter_by(status='Assigned').count()
    available = Asset.query.filter_by(status='Available').count()
    maintenance = Asset.query.filter_by(status='Maintenance').count()

    today = date.today()
    soon  = today + timedelta(days=90)
    expiring = Asset.query.filter(
        Asset.warranty_date != None,
        Asset.warranty_date <= soon,
        Asset.warranty_date >= today
    ).count()

    # Category breakdown (only assigned assets)
    cat_data = db.session.query(Asset.category, func.count(Asset.id))\
        .filter(Asset.status == 'Assigned')\
        .filter(Asset.category != None)\
        .group_by(Asset.category).all()

    # Laptop status breakdown
    laptop_total = Asset.query.filter_by(category='Laptop').count()
    laptop_available = Asset.query.filter_by(category='Laptop', status='Available').count()
    laptop_assigned = Asset.query.filter_by(category='Laptop', status='Assigned').count()
    laptop_maintenance = Asset.query.filter_by(category='Laptop', status='Maintenance').count()
    laptop_retired = Asset.query.filter_by(category='Laptop', status='Retired').count()

    return jsonify({
        'totalAssets': total,
        'assignedAssets': assigned,
        'availableAssets': available,
        'maintenanceAssets': maintenance,
        'expiringWarranties': expiring,
        'categories': [{'name': c[0], 'count': c[1]} for c in cat_data],
        'laptopStats': {
            'total': laptop_total,
            'available': laptop_available,
            'assigned': laptop_assigned,
            'maintenance': laptop_maintenance,
            'retired': laptop_retired
        }
    })

@api_bp.route('/dashboard/activity')
@cross_origin()
def api_dashboard_activity():
    logs = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).limit(20).all()
    return jsonify({'logs': [l.to_dict() for l in logs]})

@api_bp.route('/assets')
@cross_origin()
def api_assets():
    search   = request.args.get('search', '')
    category = request.args.get('category', '')
    status   = request.args.get('status', '')
    location = request.args.get('location', '')
    page     = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    query = Asset.query
    if search:
        query = query.filter(or_(
            Asset.asset_name.ilike(f'%{search}%'),
            Asset.serial_number.ilike(f'%{search}%'),
            Asset.emp_id.ilike(f'%{search}%'),
            Asset.employee_name.ilike(f'%{search}%')
        ))
    if category: query = query.filter_by(category=category)
    if status:   query = query.filter_by(status=status)
    if location: query = query.filter_by(location=location)

    sort = request.args.get('sort', 'id_asc')
    sort_map = {
        'id_asc':   Asset.id.asc(),
        'id_desc':  Asset.id.desc(),
        'emp_asc':  Asset.emp_id.asc(),
        'emp_desc': Asset.emp_id.desc(),
        'name_asc': Asset.asset_name.asc(),
    }
    paginated = query.order_by(sort_map.get(sort, Asset.id.asc())).paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        'assets': [a.to_dict() for a in paginated.items],
        'total': paginated.total,
        'pages': paginated.pages,
        'page': page
    })

@api_bp.route('/assets/<int:asset_id>')
@cross_origin()
def api_asset_detail(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    return jsonify(asset.to_dict())

@api_bp.route('/assets/<int:asset_id>', methods=['DELETE'])
@cross_origin()
def api_delete_asset(asset_id):
    """Delete an asset"""
    try:
        asset = Asset.query.get_or_404(asset_id)
        asset_name = asset.asset_name
        serial = asset.serial_number
        
        # Create audit log before deletion
        AuditService.log_asset_deleted(asset, 'admin')
        
        db.session.delete(asset)
        
        # Log activity (legacy)
        log = ActivityLog(
            user='admin',
            action='DELETE',
            module='Asset',
            description=f'Deleted asset: {asset_name} [{serial}]'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Asset "{asset_name}" deleted successfully'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
@api_bp.route('/assets/<int:asset_id>', methods=['PUT'])
@cross_origin()
def api_update_asset(asset_id):
    """Update an asset"""
    try:
        asset = Asset.query.get_or_404(asset_id)
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Track changes for audit
        changed_fields = {}
        old_status = asset.status
        old_emp_id = asset.emp_id
        old_employee_name = asset.employee_name
        
        # Parse dates helper
        def parse_date(val):
            if not val:
                return None
            try:
                return datetime.strptime(val, '%Y-%m-%d').date()
            except:
                return None
        
        # Check serial number uniqueness only if it's being changed
        if 'serial_number' in data:
            new_serial = data['serial_number'].strip()
            if new_serial != asset.serial_number:
                existing = Asset.query.filter_by(serial_number=new_serial).first()
                if existing:
                    return jsonify({'error': 'Serial number already exists'}), 409
                changed_fields['serial_number'] = (asset.serial_number, new_serial)
                asset.serial_number = new_serial
        
        # Track and update fields
        field_map = {
            'emp_id': 'emp_id',
            'employee_name': 'employee_name',
            'employee_email': 'employee_email',
            'mobile_number': 'mobile_number',
            'asset_name': 'asset_name',
            'category': 'category',
            'model_name': 'model_name',
            'os': 'os',
            'version': 'version',
            'ram': 'ram',
            'location': 'location',
            'invoice_number': 'invoice_number',
            'charger_serial': 'charger_serial',
            'old_user': 'old_user',
            'old_device': 'old_device',
            'comments': 'comments',
            'status': 'status',
        }
        
        for data_key, attr_name in field_map.items():
            if data_key in data:
                old_val = getattr(asset, attr_name)
                new_val = data[data_key]
                if str(old_val) != str(new_val):
                    changed_fields[attr_name] = (old_val, new_val)
                    setattr(asset, attr_name, new_val)
        
        # Handle email alias
        if 'email' in data:
            old_val = asset.employee_email
            new_val = data['email']
            if str(old_val) != str(new_val):
                changed_fields['employee_email'] = (old_val, new_val)
                asset.employee_email = new_val
        
        # Date fields
        if 'invoice_date' in data:
            asset.invoice_date = parse_date(data['invoice_date'])
        if 'warranty_date' in data:
            asset.warranty_date = parse_date(data['warranty_date'])
        if 'date' in data:
            asset.date = parse_date(data['date'])
        
        # Additional fields (not tracked for audit simplicity)
        if 'purchase_price' in data:
            asset.purchase_price = data['purchase_price']
        if 'quantity' in data:
            asset.quantity = data['quantity']
        if 'configuration' in data:
            asset.configuration = data['configuration']
        if 'laptop_bag_serial' in data:
            asset.laptop_bag_serial = data['laptop_bag_serial']
        if 'hard_disk_serial' in data:
            asset.hard_disk_serial = data['hard_disk_serial']
        if 'hard_disk_capacity' in data:
            asset.hard_disk_capacity = data['hard_disk_capacity']
        if 'ups_serial' in data:
            asset.ups_serial = data['ups_serial']
        if 'ups_capacity' in data:
            asset.ups_capacity = data['ups_capacity']
        if 'printer_type' in data:
            asset.printer_type = data['printer_type']
        if 'printer_model' in data:
            asset.printer_model = data['printer_model']
        if 'mobile_imei' in data:
            asset.mobile_imei = data['mobile_imei']
        if 'mobile_number_sim' in data:
            asset.mobile_number_sim = data['mobile_number_sim']
        if 'testing_status' in data:
            asset.testing_status = data['testing_status']
        
        asset.updated_at = datetime.utcnow()
        
        # Log activity (legacy)
        log = ActivityLog(
            user='admin',
            action='UPDATE',
            module='Asset',
            description=f'Updated asset: {asset.asset_name} [{asset.serial_number}]'
        )
        db.session.add(log)
        
        # Create comprehensive audit logs
        if changed_fields:
            AuditService.log_asset_updated(asset, changed_fields, 'admin')
        
        # Handle status changes
        if 'status' in changed_fields:
            new_status = changed_fields['status'][1]
            AuditService.log_status_change(asset, old_status, new_status, 'admin')
            LifecycleService.record_event(
                asset_id=asset.id,
                event_type='STATUS_CHANGED',
                from_status=old_status,
                to_status=new_status,
                performed_by='admin'
            )
        
        # Handle employee assignment changes
        new_emp_id = asset.emp_id
        if old_emp_id != new_emp_id:
            if old_emp_id and not new_emp_id:  # Returned
                AuditService.log_asset_returned(asset, old_employee_name or '', old_emp_id, 'admin', new_status=asset.status)
                LifecycleService.record_event(
                    asset_id=asset.id,
                    event_type='RETURNED',
                    from_employee_id=old_emp_id,
                    from_employee=old_employee_name,
                    from_status='Assigned',
                    to_status=asset.status,
                    performed_by='admin'
                )
            elif new_emp_id and not old_emp_id:  # Assigned
                AuditService.log_asset_assigned(asset, asset.employee_name, new_emp_id, 'admin', old_status=old_status)
                LifecycleService.record_event(
                    asset_id=asset.id,
                    event_type='ASSIGNED',
                    to_employee_id=new_emp_id,
                    to_employee=asset.employee_name,
                    from_status=old_status,
                    to_status='Assigned',
                    performed_by='admin'
                )
            elif new_emp_id and old_emp_id and new_emp_id != old_emp_id:  # Reassigned
                AuditService.log(
                    action_type='ASSET_REASSIGNED',
                    module='Asset',
                    asset_id=asset.id,
                    asset_name=asset.asset_name,
                    asset_serial=asset.serial_number,
                    category=asset.category,
                    employee_id=new_emp_id,
                    employee_name=asset.employee_name,
                    old_value=old_emp_id,
                    new_value=new_emp_id,
                    performed_by='admin',
                    remarks=f"Reassigned from {old_emp_id} to {new_emp_id}"
                )
                LifecycleService.record_event(
                    asset_id=asset.id,
                    event_type='REASSIGNED',
                    from_employee_id=old_emp_id,
                    to_employee_id=new_emp_id,
                    to_employee=asset.employee_name,
                    performed_by='admin'
                )
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'asset': asset.to_dict(),
            'message': 'Asset updated successfully'
        }), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error updating asset {asset_id}: {str(e)}')
        return jsonify({'error': str(e)}), 500

@api_bp.route('/assets', methods=['POST'])
@cross_origin()
def api_create_asset():
    """Create a new asset"""
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('asset_name') or not data.get('serial_number'):
            return jsonify({'error': 'Asset name and serial number are required'}), 400
        
        # Check duplicate serial number
        if Asset.query.filter_by(serial_number=data['serial_number'].strip()).first():
            return jsonify({'error': 'Serial number already exists'}), 409
        
        # Parse dates
        def parse_date(val):
            if not val:
                return None
            try:
                return datetime.strptime(val, '%Y-%m-%d').date()
            except:
                return None
        
        asset = Asset(
            emp_id          = data.get('emp_id', ''),
            employee_name   = data.get('employee_name', ''),
            employee_email  = data.get('employee_email') or data.get('email', ''),
            mobile_number   = data.get('mobile_number', ''),
            asset_name      = data['asset_name'].strip(),
            category        = data.get('category', ''),
            serial_number   = data['serial_number'].strip(),
            model_name      = data.get('model_name', ''),
            os              = data.get('os', ''),
            version         = data.get('version', ''),
            ram             = data.get('ram', ''),
            location        = data.get('location', ''),
            invoice_number  = data.get('invoice_number', ''),
            invoice_date    = parse_date(data.get('invoice_date')),
            warranty_date   = parse_date(data.get('warranty_date')),
            charger_serial  = data.get('charger_serial', ''),
            old_user        = data.get('old_user', ''),
            date            = parse_date(data.get('date')) or date.today(),
            old_device      = data.get('old_device', ''),
            comments        = data.get('comments', ''),
            status          = data.get('status', 'Available'),
        )
        
        db.session.add(asset)
        db.session.flush()  # Get asset.id before full commit
        
        # Log activity (legacy)
        log = ActivityLog(
            user='admin',
            action='CREATE',
            module='Asset',
            description=f'Added asset: {asset.asset_name} [{asset.serial_number}]'
        )
        db.session.add(log)
        
        # Create comprehensive audit log and lifecycle event
        AuditService.log_asset_created(asset, 'admin')
        LifecycleService.record_event(
            asset_id=asset.id,
            event_type='PROCURED',
            to_status=asset.status,
            reason='New asset added to inventory',
            performed_by='admin'
        )
        
        db.session.commit()  # Commit everything together
        
        return jsonify({
            'success': True,
            'asset': asset.to_dict(),
            'message': 'Asset created successfully'
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/auth/login', methods=['POST'])
@cross_origin()
def api_login():
    data     = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    user = User.query.filter_by(username=username).first()
    if user and check_password_hash(user.password_hash, password):
        return jsonify({
            'success': True,
            'token': f'user-{user.id}-{user.username}',
            'user': {'id': user.id, 'username': user.username, 'role': user.role}
        })
    return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

@api_bp.route('/auth/logout', methods=['POST'])
@cross_origin()
def api_logout():
    return jsonify({'success': True})

@api_bp.route('/health')
@cross_origin()
def api_health():
    return jsonify({'status': 'ok', 'message': 'API running'})

@api_bp.route('/assets/import', methods=['POST'])
@cross_origin()
def api_import_assets():
    """Import assets from Excel file"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Only Excel files (.xlsx, .xls) are allowed'}), 400
    
    try:
        import pandas as pd
        from datetime import datetime
        
        # Read Excel file
        df = pd.read_excel(file)
        
        # Expected columns
        required_cols = ['Asset Name', 'Serial Number']
        for col in required_cols:
            if col not in df.columns:
                return jsonify({'error': f'Missing required column: {col}'}), 400
        
        success_count = 0
        error_count = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # Check if serial number already exists
                if Asset.query.filter_by(serial_number=str(row.get('Serial Number', '')).strip()).first():
                    errors.append(f"Row {idx+2}: Serial number '{row.get('Serial Number')}' already exists")
                    error_count += 1
                    continue
                
                # Parse dates
                def parse_date(val):
                    if pd.isna(val) or val == '':
                        return None
                    if isinstance(val, str):
                        try:
                            return datetime.strptime(val, '%Y-%m-%d').date()
                        except:
                            try:
                                return datetime.strptime(val, '%d/%m/%Y').date()
                            except:
                                return None
                    return val.date() if hasattr(val, 'date') else None
                
                asset = Asset(
                    emp_id           = str(row.get('EMP ID', '')).strip() if not pd.isna(row.get('EMP ID')) else '',
                    employee_name    = str(row.get('Employee Name', '')).strip() if not pd.isna(row.get('Employee Name')) else '',
                    mobile_number    = str(row.get('Mobile Number', '')).strip() if not pd.isna(row.get('Mobile Number')) else '',
                    asset_name       = str(row.get('Asset Name', '')).strip(),
                    category         = str(row.get('Category', '')).strip() if not pd.isna(row.get('Category')) else '',
                    serial_number    = str(row.get('Serial Number', '')).strip(),
                    model_name       = str(row.get('Model Name', '')).strip() if not pd.isna(row.get('Model Name')) else '',
                    os               = str(row.get('OS', '')).strip() if not pd.isna(row.get('OS')) else '',
                    version          = str(row.get('Version', '')).strip() if not pd.isna(row.get('Version')) else '',
                    ram              = str(row.get('RAM', '')).strip() if not pd.isna(row.get('RAM')) else '',
                    location         = str(row.get('Location', '')).strip() if not pd.isna(row.get('Location')) else '',
                    invoice_number   = str(row.get('Invoice Number', '')).strip() if not pd.isna(row.get('Invoice Number')) else '',
                    invoice_date     = parse_date(row.get('Invoice Date')),
                    warranty_date    = parse_date(row.get('Warranty Date')),
                    charger_serial   = str(row.get('Charger Serial Number', '')).strip() if not pd.isna(row.get('Charger Serial Number')) else '',
                    old_user         = str(row.get('Old User', '')).strip() if not pd.isna(row.get('Old User')) else '',
                    date             = parse_date(row.get('Date')) or date.today(),
                    old_device       = str(row.get('Old Device', '')).strip() if not pd.isna(row.get('Old Device')) else '',
                    comments         = str(row.get('Comments', '')).strip() if not pd.isna(row.get('Comments')) else '',
                    status           = str(row.get('Status', 'Available')).strip() if not pd.isna(row.get('Status')) else 'Available'
                )
                db.session.add(asset)
                success_count += 1
            except Exception as e:
                errors.append(f"Row {idx+2}: {str(e)}")
                error_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Import completed: {success_count} assets added, {error_count} errors',
            'imported': success_count,
            'errors': error_count,
            'error_details': errors[:10]  # Return first 10 errors
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Import failed: {str(e)}'}), 500

@api_bp.route('/assets/template', methods=['GET'])
@cross_origin()
def api_download_template():
    """Download Excel template for bulk import"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Asset Import Template'
        
        # Headers
        headers = [
            'EMP ID', 'Employee Name', 'Mobile Number', 'Asset Name', 'Category',
            'Serial Number', 'Model Name', 'OS', 'Version', 'RAM', 'Location',
            'Invoice Number', 'Invoice Date', 'Warranty Date', 'Charger Serial Number',
            'Old User', 'Date', 'Old Device', 'Comments', 'Status'
        ]
        
        # Style header
        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add sample data
        sample_data = [
            'EMP001', 'John Doe', '9876543210', 'Dell Laptop XPS 15', 'Laptop',
            'SN-DELL-001', 'XPS 15 9520', 'Windows 11', '23H2', '16GB', 'HQ - Floor 2',
            'INV-2024-001', '2024-01-15', '2027-01-15', 'CHG-DELL-001',
            '', '2024-01-20', '', 'Primary work laptop', 'Assigned'
        ]
        
        for col_idx, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col_idx, value=value)
        
        # Auto-fit columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
        
        # Add instructions
        ws2 = wb.create_sheet('Instructions')
        instructions = [
            ['Asset Import Template - Instructions'],
            [''],
            ['Required Fields (must be filled):'],
            ['  • Asset Name'],
            ['  • Serial Number (must be unique)'],
            [''],
            ['Optional Fields:'],
            ['  • All other fields are optional'],
            [''],
            ['Date Format:'],
            ['  • Use YYYY-MM-DD format (e.g., 2024-01-15)'],
            ['  • Or DD/MM/YYYY format (e.g., 15/01/2024)'],
            [''],
            ['Status Options:'],
            ['  • Available'],
            ['  • Assigned'],
            ['  • Maintenance'],
            [''],
            ['Notes:'],
            ['  • Delete the sample row before importing your data'],
            ['  • Serial numbers must be unique'],
            ['  • Maximum 1000 rows per import'],
        ]
        
        for row_idx, row_data in enumerate(instructions, 1):
            ws2.cell(row=row_idx, column=1, value=row_data[0])
        
        ws2.column_dimensions['A'].width = 60
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Asset_Import_Template.xlsx'
        )
    except Exception as e:
        return jsonify({'error': f'Failed to generate template: {str(e)}'}), 500

# ── USER MANAGEMENT API ───────────────────────────────────────────────────────
@api_bp.route('/users', methods=['GET'])
@cross_origin()
def api_get_users():
    users = User.query.all()
    return jsonify([{
        'id': u.id,
        'username': u.username,
        'email': u.email,
        'role': u.role,
        'created_at': u.created_at.isoformat() if hasattr(u, 'created_at') and u.created_at else None
    } for u in users])

@api_bp.route('/users', methods=['POST'])
@cross_origin()
def api_create_user():
    data = request.get_json()
    username = data.get('username', '').strip()
    email    = data.get('email', '').strip()
    password = data.get('password', '')
    role     = data.get('role', 'standard')

    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'Username already exists'}), 400
    if email and User.query.filter_by(email=email).first():
        return jsonify({'error': 'Email already exists'}), 400
    if not email:
        email = None

    user = User(
        username=username,
        email=email,
        password_hash=generate_password_hash(password),
        role=role
    )
    db.session.add(user)
    db.session.commit()
    return jsonify({'message': 'User created', 'id': user.id}), 201

@api_bp.route('/users/<int:uid>', methods=['PUT'])
@cross_origin()
def api_update_user(uid):
    user = User.query.get_or_404(uid)
    data = request.get_json()

    if 'username' in data and data['username'].strip():
        existing = User.query.filter_by(username=data['username'].strip()).first()
        if existing and existing.id != uid:
            return jsonify({'error': 'Username already exists'}), 400
        user.username = data['username'].strip()

    if 'email' in data:
        user.email = data['email'].strip()
    if 'role' in data:
        user.role = data['role']
    if 'password' in data and data['password']:
        user.password_hash = generate_password_hash(data['password'])

    db.session.commit()
    return jsonify({'message': 'User updated'})

@api_bp.route('/users/<int:uid>', methods=['DELETE'])
@cross_origin()
def api_delete_user(uid):
    user = User.query.get_or_404(uid)
    if user.username == 'admin':
        return jsonify({'error': 'Cannot delete the main admin user'}), 403
    db.session.delete(user)
    db.session.commit()
    return jsonify({'message': 'User deleted'})


# ── EMAIL API ─────────────────────────────────────────────────────────────────
@api_bp.route('/users/<int:uid>/smtp-password', methods=['PUT'])
@cross_origin()
def api_update_smtp_password(uid):
    user = User.query.get_or_404(uid)
    data = request.get_json()
    pwd  = data.get('smtp_password', '').strip()
    if not pwd:
        return jsonify({'error': 'Password is required'}), 400
    from werkzeug.security import generate_password_hash
    user.smtp_password = pwd          # store plain for SMTP use
    db.session.commit()
    return jsonify({'message': 'SMTP password saved'})


@api_bp.route('/assets/<int:asset_id>/send-assignment-email', methods=['POST'])
@cross_origin()
def api_send_assignment_email(asset_id):
    try:
        data             = request.get_json()
        recipient_email  = data.get('recipient_email', '').strip()
        sender_user_id   = data.get('sender_user_id')

        if not recipient_email:
            return jsonify({'error': 'Recipient email is required'}), 400

        asset = Asset.query.get_or_404(asset_id)
        sender = User.query.get(sender_user_id) if sender_user_id else None

        if not sender or not sender.email:
            return jsonify({'error': 'Sender email not configured'}), 400
        if not sender.smtp_password:
            return jsonify({'error': 'SMTP password not set. Please update it in Settings.'}), 400

        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        subject = f"Asset Assignment Acknowledgement – {asset.asset_name}"

        html_body = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;">
          <div style="background:#1e3a5f;padding:24px;text-align:center;">
            <h2 style="color:#fff;margin:0;">Tectoro Asset Management</h2>
            <p style="color:#93c5fd;margin:4px 0 0;">Asset Assignment Confirmation</p>
          </div>
          <div style="padding:32px;">
            <p style="font-size:16px;">Dear <strong>{asset.employee_name or recipient_email}</strong>,</p>
            <p>This is to confirm that the following asset has been assigned to you:</p>

            <table style="width:100%;border-collapse:collapse;margin:20px 0;">
              <tr style="background:#f1f5f9;">
                <td style="padding:10px 14px;font-weight:bold;width:40%;">Asset Name</td>
                <td style="padding:10px 14px;">{asset.asset_name}</td>
              </tr>
              <tr>
                <td style="padding:10px 14px;font-weight:bold;">Category</td>
                <td style="padding:10px 14px;">{asset.category or '—'}</td>
              </tr>
              <tr style="background:#f1f5f9;">
                <td style="padding:10px 14px;font-weight:bold;">Serial Number</td>
                <td style="padding:10px 14px;">{asset.serial_number}</td>
              </tr>
              <tr>
                <td style="padding:10px 14px;font-weight:bold;">Model</td>
                <td style="padding:10px 14px;">{asset.model_name or '—'}</td>
              </tr>
              <tr style="background:#f1f5f9;">
                <td style="padding:10px 14px;font-weight:bold;">Assignment Date</td>
                <td style="padding:10px 14px;">{asset.date or '—'}</td>
              </tr>
              <tr>
                <td style="padding:10px 14px;font-weight:bold;">Location</td>
                <td style="padding:10px 14px;">{asset.location or '—'}</td>
              </tr>
              <tr style="background:#f1f5f9;">
                <td style="padding:10px 14px;font-weight:bold;">EMP ID</td>
                <td style="padding:10px 14px;">{asset.emp_id or '—'}</td>
              </tr>
            </table>

            <p style="background:#fef3c7;border-left:4px solid #f59e0b;padding:12px 16px;border-radius:4px;">
              Please acknowledge receipt of this asset by replying to this email.
              You are responsible for the safe custody of this asset.
            </p>

            <p style="margin-top:24px;">Regards,<br/>
            <strong>{sender.username}</strong><br/>
            Tectoro Asset Management Team</p>
          </div>
          <div style="background:#f8fafc;padding:16px;text-align:center;font-size:12px;color:#94a3b8;">
            This is an automated message from Tectoro Asset Management System.
          </div>
        </div>
        """

        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = sender.email
        msg['To']      = recipient_email
        msg.attach(MIMEText(html_body, 'html'))

        with smtplib.SMTP('smtp.office365.com', 587) as server:
            server.starttls()
            server.login(sender.email, sender.smtp_password)
            server.sendmail(sender.email, recipient_email, msg.as_string())

        log = ActivityLog(
            user=sender.username,
            action='EMAIL',
            module='Asset',
            description=f'Assignment email sent to {recipient_email} for {asset.asset_name} [{asset.serial_number}]'
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({'success': True, 'message': f'Email sent to {recipient_email}'})

    except smtplib.SMTPAuthenticationError:
        return jsonify({'error': 'SMTP authentication failed. Check your email/password in Settings.'}), 401
    except smtplib.SMTPException as e:
        return jsonify({'error': f'Email sending failed: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── REPORTS API ───────────────────────────────────────────────────────────────
@api_bp.route('/reports/activity')
@cross_origin()
def api_reports_activity():
    page     = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    logs     = ActivityLog.query.order_by(ActivityLog.timestamp.desc())\
                   .paginate(page=page, per_page=per_page, error_out=False)
    return jsonify({
        'logs':  [l.to_dict() for l in logs.items],
        'total': logs.total,
        'pages': logs.pages,
        'page':  page
    })

@api_bp.route('/reports/export/csv')
@cross_origin()
def api_reports_export_csv():
    assets = Asset.query.all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Sl No','EMP ID','Employee Name','Mobile','Employee Email','Asset Name',
                     'Category','Serial Number','Model','OS','Version','RAM','Location',
                     'Invoice Number','Invoice Date','Warranty Date','Charger Serial',
                     'Old User','Date','Old Device','Comments','Status'])
    for a in assets:
        writer.writerow([
            a.id, a.emp_id or '', a.employee_name or '', a.mobile_number or '',
            a.employee_email or '', a.asset_name, a.category or '', a.serial_number,
            a.model_name or '', a.os or '', a.version or '', a.ram or '', a.location or '',
            a.invoice_number or '', a.invoice_date or '', a.warranty_date or '',
            a.charger_serial or '', a.old_user or '', a.date or '',
            a.old_device or '', a.comments or '', a.status
        ])
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'assets_{date.today()}.csv'
    )

@api_bp.route('/reports/export/excel')
@cross_origin()
def api_reports_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        assets = Asset.query.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Assets'
        headers = ['Sl No','EMP ID','Employee Name','Mobile','Employee Email','Asset Name',
                   'Category','Serial Number','Model','OS','Version','RAM','Location',
                   'Invoice Number','Invoice Date','Warranty Date','Status']
        hfill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        hfont = Font(color='FFFFFF', bold=True)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal='center')
        for ri, a in enumerate(assets, 2):
            row = [a.id, a.emp_id or '', a.employee_name or '', a.mobile_number or '',
                   a.employee_email or '', a.asset_name, a.category or '', a.serial_number,
                   a.model_name or '', a.os or '', a.version or '', a.ram or '', a.location or '',
                   a.invoice_number or '', str(a.invoice_date or ''), str(a.warranty_date or ''), a.status]
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        for ci in range(1, len(headers)+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'assets_{date.today()}.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ACKNOWLEDGMENT & EMAIL CONFIG ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

# ── Send acknowledgment email ─────────────────────────────────────────────────
@api_bp.route('/assets/<int:asset_id>/send-ack-email', methods=['POST'])
@require_not_viewer
def send_ack_email(asset_id):
    from email_service import send_acknowledgment_email
    asset = Asset.query.get_or_404(asset_id)

    if not asset.employee_email:
        return jsonify({'success': False, 'error': 'No employee email on this asset'}), 400
    if asset.ack_status == 'Acknowledged':
        return jsonify({'success': False, 'error': 'Asset already acknowledged'}), 400

    # Resolve sender name from Bearer token
    auth_header  = request.headers.get('Authorization', '')
    token_val    = auth_header.replace('Bearer ', '')
    assigned_by  = 'IT Admin'
    if token_val.startswith('user-'):
        parts = token_val.split('-')
        if len(parts) >= 3:
            assigned_by = parts[2]

    base_url = request.host_url.rstrip('/')
    # Prefer env var for production
    import os
    base_url = os.environ.get('APP_BASE_URL', base_url)

    success, error = send_acknowledgment_email(asset, assigned_by, base_url)
    if success:
        log_activity('EMAIL', 'Asset',
                     f'Ack email sent to {asset.employee_email} for {asset.asset_name}')
        return jsonify({'success': True,
                        'message': f'Acknowledgment email sent to {asset.employee_email}'})
    return jsonify({'success': False, 'error': error}), 500


# ── Employee clicks ack link (public, no auth) ────────────────────────────────
@api_bp.route('/acknowledge/<string:token>', methods=['GET'])
def acknowledge_asset(token):
    from email_service import get_ack_success_html
    asset = Asset.query.filter_by(ack_token=token).first()

    if not asset:
        return ('''<!DOCTYPE html><html><body
            style="font-family:Arial;text-align:center;padding:60px;background:#f1f5f9;">
            <div style="max-width:420px;margin:0 auto;background:#fff;padding:48px;
                        border-radius:12px;box-shadow:0 4px 16px rgba(0,0,0,.1);">
            <h2 style="color:#dc2626;">&#10007; Invalid Link</h2>
            <p style="color:#475569;">This link is invalid or has already been used.</p>
            </div></body></html>''', 404)

    if asset.ack_status == 'Acknowledged':
        acked_on = asset.ack_received_at.strftime('%B %d, %Y at %H:%M UTC') if asset.ack_received_at else '—'
        return (f'''<!DOCTYPE html><html><body
            style="font-family:Arial;text-align:center;padding:60px;background:#f1f5f9;">
            <div style="max-width:420px;margin:0 auto;background:#fff;padding:48px;
                        border-radius:12px;box-shadow:0 4px 16px rgba(0,0,0,.1);">
            <h2 style="color:#2563eb;">Already Acknowledged</h2>
            <p style="color:#475569;">You confirmed receipt of
               <strong>{asset.asset_name}</strong> on {acked_on}.</p>
            </div></body></html>''')

    now = datetime.utcnow()
    if asset.ack_expires_at and now > asset.ack_expires_at:
        return ('''<!DOCTYPE html><html><body
            style="font-family:Arial;text-align:center;padding:60px;background:#f1f5f9;">
            <div style="max-width:420px;margin:0 auto;background:#fff;padding:48px;
                        border-radius:12px;box-shadow:0 4px 16px rgba(0,0,0,.1);">
            <h2 style="color:#f59e0b;">&#9888; Link Expired</h2>
            <p style="color:#475569;">Please contact IT support to resend the acknowledgment.</p>
            </div></body></html>''', 410)

    asset.ack_status     = 'Acknowledged'
    asset.ack_received_at = now
    asset.ack_token      = None
    db.session.commit()

    log_activity('ACKNOWLEDGE', 'Asset',
                 f'{asset.employee_name} acknowledged {asset.asset_name} (SN:{asset.serial_number})')
    return get_ack_success_html(asset)


# ── Get ack status ────────────────────────────────────────────────────────────
@api_bp.route('/assets/<int:asset_id>/ack-status', methods=['GET'])
def get_ack_status(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    return jsonify({
        'asset_id':        asset.id,
        'ack_status':      asset.ack_status or 'Not Sent',
        'ack_sent_at':     asset.ack_sent_at.isoformat() if asset.ack_sent_at else None,
        'ack_received_at': asset.ack_received_at.isoformat() if asset.ack_received_at else None,
        'ack_expires_at':  asset.ack_expires_at.isoformat() if asset.ack_expires_at else None,
        'ack_sent_by':     asset.ack_sent_by or '',
    })


# ── Email Config CRUD ─────────────────────────────────────────────────────────
@api_bp.route('/email-config', methods=['GET'])
def get_email_config():
    from models import EmailConfig
    cfg = EmailConfig.query.filter_by(is_active=True).order_by(EmailConfig.id.desc()).first()
    if not cfg:
        return jsonify({'configured': False})
    return jsonify({'configured': True, 'config': cfg.to_dict()})


@api_bp.route('/email-config', methods=['POST'])
@require_role('admin')
def save_email_config():
    from models import EmailConfig
    from email_service import encrypt_password
    data = request.get_json() or {}

    required = ['sender_email', 'sender_name', 'smtp_server', 'smtp_port',
                 'smtp_username', 'smtp_password']
    for field in required:
        if not data.get(field):
            return jsonify({'success': False, 'error': f'{field} is required'}), 400

    # Deactivate old configs
    EmailConfig.query.update({'is_active': False})

    cfg = EmailConfig(
        sender_email      = data['sender_email'].strip(),
        sender_name       = data.get('sender_name', 'IT Asset Management').strip(),
        smtp_server       = data.get('smtp_server', 'smtp.office365.com').strip(),
        smtp_port         = int(data.get('smtp_port', 587)),
        smtp_username     = data['smtp_username'].strip(),
        smtp_password_enc = encrypt_password(data['smtp_password']),
        use_tls           = bool(data.get('use_tls', True)),
        is_active         = True,
        created_by        = data.get('created_by', 'admin'),
    )
    db.session.add(cfg)
    db.session.commit()

    log_activity('UPDATE', 'EmailConfig',
                 f'Email config updated by admin — sender: {cfg.sender_email}')
    return jsonify({'success': True, 'config': cfg.to_dict()})


@api_bp.route('/email-config/test', methods=['POST'])
@require_role('admin')
def test_email_config():
    from email_service import test_smtp_config, decrypt_password
    from models import EmailConfig
    data = request.get_json() or {}

    if not data.get('test_recipient'):
        return jsonify({'success': False, 'error': 'test_recipient is required'}), 400

    # Load saved config and fill missing fields
    cfg = EmailConfig.query.filter_by(is_active=True).first()
    if not cfg:
        return jsonify({'success': False, 'error': 'No email config saved yet'}), 400

    smtp_server   = data.get('smtp_server')   or cfg.smtp_server
    smtp_port     = int(data.get('smtp_port') or cfg.smtp_port)
    smtp_username = data.get('smtp_username') or cfg.smtp_username
    sender_email  = data.get('sender_email')  or cfg.sender_email
    use_tls       = data.get('use_tls', cfg.use_tls)
    password      = data.get('smtp_password') or decrypt_password(cfg.smtp_password_enc)

    success, error = test_smtp_config(
        smtp_server    = smtp_server,
        smtp_port      = smtp_port,
        smtp_username  = smtp_username,
        plain_password = password,
        use_tls        = use_tls,
        sender_email   = sender_email,
        test_recipient = data['test_recipient'],
    )

    if success:
        # Update test timestamp if a config exists
        from models import EmailConfig
        cfg = EmailConfig.query.filter_by(is_active=True).first()
        if cfg:
            cfg.last_tested_at   = datetime.utcnow()
            cfg.last_test_status = 'success'
            db.session.commit()
        return jsonify({'success': True, 'message': f'Test email sent to {data["test_recipient"]}'})

    # Record failure
    from models import EmailConfig
    cfg = EmailConfig.query.filter_by(is_active=True).first()
    if cfg:
        cfg.last_tested_at   = datetime.utcnow()
        cfg.last_test_status = 'failed'
        db.session.commit()
    return jsonify({'success': False, 'error': error}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@api_bp.route('/employees', methods=['GET'])
def get_employees():
    from models import Employee
    q = request.args.get('q', '').strip()
    # Filter for active employees (including NULL is_active which means active)
    query = Employee.query.filter(
        or_(Employee.is_active == True, Employee.is_active == None)
    )
    if q:
        query = query.filter(
            or_(Employee.emp_id.ilike(f'%{q}%'),
                Employee.employee_name.ilike(f'%{q}%'),
                Employee.email.ilike(f'%{q}%'))
        )
    employees = query.order_by(Employee.employee_name).limit(20).all()
    return jsonify([e.to_dict() for e in employees])

@api_bp.route('/employees/<string:emp_id>', methods=['GET'])
def get_employee(emp_id):
    from models import Employee
    emp = Employee.query.filter_by(emp_id=emp_id).first()
    if not emp:
        return jsonify({'found': False})
    return jsonify({'found': True, 'employee': emp.to_dict()})

@api_bp.route('/employees', methods=['POST'])
@require_not_viewer
def create_or_update_employee():
    from models import Employee
    data = request.get_json() or {}
    if not data.get('emp_id') or not data.get('employee_name'):
        return jsonify({'error': 'emp_id and employee_name are required'}), 400

    emp = Employee.query.filter_by(emp_id=data['emp_id']).first()
    if emp:
        emp.employee_name = data.get('employee_name', emp.employee_name)
        emp.email         = data.get('email', emp.email)
        emp.mobile_number = data.get('mobile_number', emp.mobile_number)
        emp.department    = data.get('department', emp.department)
        emp.location      = data.get('location', emp.location)
    else:
        emp = Employee(
            emp_id        = data['emp_id'],
            employee_name = data['employee_name'],
            email         = data.get('email', ''),
            mobile_number = data.get('mobile_number', ''),
            department    = data.get('department', ''),
            location      = data.get('location', ''),
        )
        db.session.add(emp)
    db.session.commit()
    return jsonify({'success': True, 'employee': emp.to_dict()})


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN PROFILE ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@api_bp.route('/admin-profile', methods=['GET'])
def get_admin_profile():
    from models import AdminProfile
    profile = AdminProfile.query.first()
    if not profile:
        return jsonify({'configured': False})
    return jsonify({'configured': True, 'profile': profile.to_dict()})

@api_bp.route('/admin-profile', methods=['POST'])
@require_role('admin')
def save_admin_profile():
    from models import AdminProfile
    data = request.get_json() or {}
    profile = AdminProfile.query.first()
    if profile:
        profile.name        = data.get('name', profile.name)
        profile.email       = data.get('email', profile.email)
        profile.phone       = data.get('phone', profile.phone)
        profile.department  = data.get('department', profile.department)
        profile.designation = data.get('designation', profile.designation)
        profile.updated_at  = datetime.utcnow()
    else:
        profile = AdminProfile(
            name        = data.get('name', ''),
            email       = data.get('email', ''),
            phone       = data.get('phone', ''),
            department  = data.get('department', ''),
            designation = data.get('designation', ''),
        )
        db.session.add(profile)
    db.session.commit()
    log_activity('UPDATE', 'AdminProfile', f'Admin profile updated: {profile.name}')
    return jsonify({'success': True, 'profile': profile.to_dict()})
