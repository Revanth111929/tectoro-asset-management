#!/usr/bin/env python3
"""
api_server.py
Flask REST API backend for the IT Asset Management React frontend.
Handles all CRUD operations, authentication, and report exports.
SECURITY: JWT authentication, rate limiting, CORS restrictions
"""

import os, csv, io, logging
from datetime import datetime, date, timedelta
from flask import Flask, jsonify, request, send_file, send_from_directory, make_response
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import or_, func
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# ── App setup ─────────────────────────────────────────────────────────────────
basedir = os.path.abspath(os.path.dirname(__file__))
build_dir = os.path.join(basedir, 'frontend', 'build')

app = Flask(__name__, static_folder=build_dir, static_url_path='')

# Secure configuration from environment
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', os.urandom(32).hex())

# Database selection: strictly isolated per environment (office vs render demo).
# See db_config.py — this fails startup rather than silently falling back.
from db_config import resolve_database_uri, is_render_env, DatabaseConfigError
try:
    _db_uri, APP_ENV = resolve_database_uri(basedir)
except DatabaseConfigError as exc:
    raise SystemExit(str(exc))
app.config['SQLALCHEMY_DATABASE_URI'] = _db_uri

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 10,
    'pool_recycle': 3600,
    'pool_pre_ping': True
}

# Secure CORS configuration
allowed_origins = os.getenv('ALLOWED_ORIGINS', 'http://localhost:3000,http://192.168.20.180:3000').split(',')
CORS(app, resources={
    r"/api/*": {
        "origins": allowed_origins,
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "expose_headers": ["Content-Range", "X-Content-Range"],
        "supports_credentials": True,
        "max_age": 3600
    }
})

# Configure logging
log_level = os.getenv('LOG_LEVEL', 'INFO')
logging.basicConfig(
    level=getattr(logging, log_level),
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(basedir, 'logs', 'app.log')),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Import models AFTER app is configured
from models import db, Asset, ActivityLog, User, Employee, Onboarding, OnboardingAssetAssignment
from services.audit_service import AuditService, LifecycleService
from utils.auth import generate_access_token, generate_refresh_token, token_required, admin_required, get_current_user, non_viewer_required
from utils.rate_limit import init_limiter, limit_login, limit_api, limit_expensive

db.init_app(app)

# ── Unmissable startup trace (Production Architecture Audit, Step 6) ─────────
# Plain print() to stdout/stderr's captured console — cannot be silenced by
# LOG_LEVEL, a logger config, or --quiet flags. This is intentional: whoever
# is watching this process's console (locally or in Render's Logs tab) must
# always be able to see exactly which database this process is talking to.
with app.app_context():
    _engine_url = str(db.engine.url)
print("=" * 60, flush=True)
print("Backend:            api_server.py", flush=True)
print(f"Environment:        {APP_ENV}", flush=True)
print(f"Database URI:       {_db_uri}", flush=True)
print(f"Resolved DB file:   {_db_uri.replace('sqlite:///', '') if _db_uri.startswith('sqlite') else '(non-sqlite backend)'}", flush=True)
print(f"SQLAlchemy Engine:  {_engine_url}", flush=True)
print(f"PID:                {os.getpid()}", flush=True)
print("=" * 60, flush=True)

# Initialize rate limiter
limiter = init_limiter(app)

# Security Headers Middleware
@app.after_request
def add_security_headers(response):
    """Add security headers to all responses"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# NOTE: NOT registering routes.py blueprints as they are template-based, not API
# All API routes are defined directly in this file

# ── Helpers ───────────────────────────────────────────────────────────────────
def parse_date(val):
    """Parse YYYY-MM-DD string to date object"""
    if val:
        try:
            return datetime.strptime(val, '%Y-%m-%d').date()
        except ValueError:
            logger.warning(f"Invalid date format: {val}")
    return None

def safe_float(val):
    """Helper to safely convert to float, returning None for empty strings"""
    if val == '' or val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        logger.warning(f"Invalid float value: {val}")
        return None

def safe_int(val, default=None):
    """Helper to safely convert to int, returning default for empty strings"""
    if val == '' or val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        logger.warning(f"Invalid int value: {val}")
        return default

def log_activity(action, module, description, user='system'):
    """Legacy activity logging - kept for compatibility"""
    try:
        # Handle both string and dict user parameter
        if isinstance(user, dict):
            user = user.get('username', 'system')
        entry = ActivityLog(user=user, action=action, module=module, description=description)
        db.session.add(entry)
        # caller must commit
    except Exception as e:
        logger.error(f"Error logging activity: {e}")

# ── Seed sample data ──────────────────────────────────────────────────────────
def seed_data():
    if User.query.first():
        return  # already seeded

    print("🌱 Seeding sample data...")

    # Admin user  (password: admin123)
    admin = User(
        username='admin',
        email='admin@company.com',
        password_hash=generate_password_hash('admin123'),
        role='admin'
    )
    db.session.add(admin)

    # Sample assets with all 20 columns
    samples = [
        Asset(
            emp_id='EMP001', employee_name='Alice Johnson', mobile_number='9876543210',
            asset_name='Dell Laptop XPS 15', category='Laptop',
            serial_number='SN-DELL-001', model_name='XPS 15 9500',
            os='Windows 11', version='23H2', ram='16GB',
            location='HQ - Floor 2', invoice_number='INV-2023-001',
            invoice_date=date(2023, 1, 15), warranty_date=date(2026, 1, 15),
            charger_serial='CHG-DELL-001', old_user='', date=date(2023, 1, 20),
            old_device='', comments='Primary work laptop', status='Assigned'
        ),
        Asset(
            emp_id='EMP002', employee_name='Bob Williams', mobile_number='9876543211',
            asset_name='HP EliteBook 840', category='Laptop',
            serial_number='SN-HP-002', model_name='EliteBook 840 G9',
            os='Windows 11', version='22H2', ram='8GB',
            location='HQ - Floor 1', invoice_number='INV-2023-002',
            invoice_date=date(2023, 3, 10), warranty_date=date(2026, 3, 10),
            charger_serial='CHG-HP-002', old_user='Carol Davis', date=date(2023, 3, 15),
            old_device='HP EliteBook 830', comments='Transferred from Carol', status='Assigned'
        ),
        Asset(
            emp_id='', employee_name='', mobile_number='',
            asset_name='Apple MacBook Pro 14"', category='Laptop',
            serial_number='SN-APL-003', model_name='MacBook Pro M2',
            os='macOS', version='Ventura 13.5', ram='16GB',
            location='Store Room', invoice_number='INV-2023-003',
            invoice_date=date(2023, 5, 5), warranty_date=date(2026, 5, 5),
            charger_serial='CHG-APL-003', old_user='', date=date(2023, 5, 10),
            old_device='', comments='Available for assignment', status='Available'
        ),
        Asset(
            emp_id='EMP003', employee_name='Carol Davis', mobile_number='9876543212',
            asset_name='Lenovo ThinkPad X1', category='Laptop',
            serial_number='SN-LEN-004', model_name='ThinkPad X1 Carbon Gen 11',
            os='Ubuntu', version='22.04 LTS', ram='16GB',
            location='HQ - Floor 3', invoice_number='INV-2023-004',
            invoice_date=date(2023, 7, 12), warranty_date=date(2026, 7, 12),
            charger_serial='CHG-LEN-004', old_user='', date=date(2023, 7, 15),
            old_device='', comments='Dev machine', status='Assigned'
        ),
        Asset(
            emp_id='EMP004', employee_name='David Brown', mobile_number='9876543213',
            asset_name='Dell Latitude 5540', category='Laptop',
            serial_number='SN-DELL-005', model_name='Latitude 5540',
            os='Windows 10', version='22H2', ram='8GB',
            location='Branch Office', invoice_number='INV-2022-005',
            invoice_date=date(2022, 9, 1), warranty_date=date(2025, 9, 1),
            charger_serial='CHG-DELL-005', old_user='Eva Martinez', date=date(2022, 9, 5),
            old_device='Dell Latitude 5430', comments='Warranty expiring soon', status='Assigned'
        ),
        Asset(
            emp_id='', employee_name='', mobile_number='',
            asset_name='HP ProBook 450', category='Laptop',
            serial_number='SN-HP-006', model_name='ProBook 450 G10',
            os='Windows 11', version='23H2', ram='8GB',
            location='Store Room', invoice_number='INV-2024-006',
            invoice_date=date(2024, 1, 20), warranty_date=date(2027, 1, 20),
            charger_serial='CHG-HP-006', old_user='', date=date(2024, 1, 25),
            old_device='', comments='New stock', status='Available'
        ),
        Asset(
            emp_id='EMP005', employee_name='Eva Martinez', mobile_number='9876543214',
            asset_name='Asus ZenBook 14', category='Laptop',
            serial_number='SN-ASUS-007', model_name='ZenBook 14 OLED',
            os='Windows 11', version='23H2', ram='16GB',
            location='HQ - Floor 2', invoice_number='INV-2023-007',
            invoice_date=date(2023, 11, 10), warranty_date=date(2026, 11, 10),
            charger_serial='CHG-ASUS-007', old_user='', date=date(2023, 11, 15),
            old_device='', comments='Design team laptop', status='Assigned'
        ),
        Asset(
            emp_id='', employee_name='', mobile_number='',
            asset_name='Lenovo IdeaPad 3', category='Laptop',
            serial_number='SN-LEN-008', model_name='IdeaPad 3 Gen 8',
            os='Windows 11', version='22H2', ram='4GB',
            location='Maintenance', invoice_number='INV-2022-008',
            invoice_date=date(2022, 4, 8), warranty_date=date(2025, 4, 8),
            charger_serial='CHG-LEN-008', old_user='Bob Williams', date=date(2022, 4, 10),
            old_device='', comments='Under repair - keyboard issue', status='Maintenance'
        ),
    ]

    for s in samples:
        db.session.add(s)

    # Activity logs
    logs = [
        ActivityLog(user='admin', action='CREATE', module='Asset', description='Added Dell Laptop XPS 15 [SN-DELL-001]'),
        ActivityLog(user='admin', action='ASSIGN', module='Asset', description='Assigned HP EliteBook 840 to Bob Williams'),
        ActivityLog(user='admin', action='CREATE', module='Asset', description='Added Apple MacBook Pro 14"'),
        ActivityLog(user='admin', action='UPDATE', module='Asset', description='Updated Lenovo IdeaPad 3 status to Maintenance'),
    ]
    db.session.add_all(logs)
    db.session.commit()
    print("✅ Sample data seeded!")

# ══════════════════════════════════════════════════════════════════════════════
# AUTH ROUTES (JWT-based secure authentication)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/auth/login', methods=['POST'])
@limiter.limit(limit_login())
def login():
    """Authenticate user and return JWT tokens"""
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')

    if not username or not password:
        logger.warning(f"Login attempt with missing credentials from {request.remote_addr}")
        return jsonify({'error': 'Username and password are required'}), 400

    user = User.query.filter_by(username=username).first()
    if user and check_password_hash(user.password_hash, password):
        # Generate secure JWT tokens
        access_token = generate_access_token(user.id, user.username, user.role)
        refresh_token = generate_refresh_token(user.id, user.username)
        
        logger.info(f"Successful login: {username} from {request.remote_addr}")
        
        return jsonify({
            'success': True,
            'access_token': access_token,
            'refresh_token': refresh_token,
            'token': access_token,  # For backward compatibility
            'token_type': 'Bearer',
            'expires_in': int(os.getenv('JWT_ACCESS_TOKEN_EXPIRES', 3600)),
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'role': user.role
            }
        }), 200

    logger.warning(f"Failed login attempt for username: {username} from {request.remote_addr}")
    return jsonify({'error': 'Invalid username or password'}), 401

@app.route('/api/auth/logout', methods=['POST'])
@token_required
def logout():
    """Logout user (client should discard tokens)"""
    user = get_current_user()
    if user:
        logger.info(f"User logged out: {user.get('username')}")
    return jsonify({'success': True, 'message': 'Logged out successfully'}), 200

@app.route('/api/auth/refresh', methods=['POST'])
@limiter.limit("10 per minute")
def refresh_token():
    """Refresh access token using refresh token"""
    from utils.auth import decode_token
    
    data = request.get_json() or {}
    refresh_token_str = data.get('refresh_token', '')
    
    if not refresh_token_str:
        return jsonify({'error': 'Refresh token is required'}), 400
    
    payload = decode_token(refresh_token_str)
    if 'error' in payload:
        return jsonify({'error': payload['error']}), 401
    
    if payload.get('type') != 'refresh':
        return jsonify({'error': 'Invalid token type'}), 401
    
    # Generate new access token
    user = User.query.get(payload.get('user_id'))
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    new_access_token = generate_access_token(user.id, user.username, user.role)
    
    return jsonify({
        'success': True,
        'access_token': new_access_token,
        'token': new_access_token,
        'token_type': 'Bearer',
        'expires_in': int(os.getenv('JWT_ACCESS_TOKEN_EXPIRES', 3600))
    }), 200

@app.route('/api/auth/me', methods=['GET'])
@token_required
def get_current_user_info():
    """Get current authenticated user info"""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Not authenticated'}), 401
    
    user_obj = User.query.get(user.get('id'))
    if not user_obj:
        return jsonify({'error': 'User not found'}), 404
    
    return jsonify({
        'id': user_obj.id,
        'username': user_obj.username,
        'email': user_obj.email,
        'role': user_obj.role,
        'created_at': user_obj.created_at.isoformat() if user_obj.created_at else None
    }), 200

# ══════════════════════════════════════════════════════════════════════════════
# USER MANAGEMENT (Admin Users)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    """Get all admin users"""
    users = User.query.all()
    return jsonify([{
        'id': u.id,
        'username': u.username,
        'email': u.email,
        'role': u.role,
        'created_at': u.created_at.isoformat() if hasattr(u, 'created_at') and u.created_at else None
    } for u in users]), 200

@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    """Create new admin user"""
    data = request.get_json() or {}
    
    username = data.get('username', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '')
    role = data.get('role', 'user')
    
    # Convert empty email to None to avoid unique constraint violation
    if not email:
        email = None
    
    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    
    # Password strength validation
    if len(password) < 8:
        return jsonify({'error': 'Password must be at least 8 characters long'}), 400
    
    # Validate role
    valid_roles = ['admin', 'user', 'viewer']
    if role not in valid_roles:
        return jsonify({'error': f'Invalid role. Must be one of: {", ".join(valid_roles)}'}), 400
    
    # Check if user exists
    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'Username already exists'}), 409
    
    # Check if email exists (only if email is provided)
    if email and User.query.filter_by(email=email).first():
        return jsonify({'error': 'Email already exists'}), 409
    
    user = User(
        username=username,
        email=email,
        password_hash=generate_password_hash(password),
        role=role
    )
    db.session.add(user)
    
    current_user = get_current_user()
    log_activity('CREATE', 'User', f'Created user: {username} with role {role}', current_user.get('username') if current_user else 'system')
    db.session.commit()
    
    logger.info(f"New user created: {username} (role: {role}) by {current_user.get('username') if current_user else 'system'}")
    
    return jsonify({'success': True, 'user': {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'role': user.role
    }}), 201

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def update_user(user_id):
    """Update admin user"""
    user = User.query.get_or_404(user_id)
    data = request.get_json() or {}
    
    if 'email' in data:
        email = data['email'].strip() if data['email'] else None
        # Check if email is being changed and already exists
        if email and email != user.email:
            existing = User.query.filter_by(email=email).first()
            if existing and existing.id != user_id:
                return jsonify({'error': 'Email already exists'}), 409
        user.email = email
    
    if 'role' in data:
        role = data['role']
        valid_roles = ['admin', 'user', 'viewer']
        if role not in valid_roles:
            return jsonify({'error': f'Invalid role. Must be one of: {", ".join(valid_roles)}'}), 400
        user.role = role
    
    if 'password' in data and data['password']:
        if len(data['password']) < 8:
            return jsonify({'error': 'Password must be at least 8 characters long'}), 400
        user.password_hash = generate_password_hash(data['password'])
    
    current_user = get_current_user()
    log_activity('UPDATE', 'User', f'Updated user: {user.username} (role: {user.role})', current_user.get('username') if current_user else 'system')
    db.session.commit()
    
    logger.info(f"User updated: {user.username} (role: {user.role}) by {current_user.get('username') if current_user else 'system'}")
    
    return jsonify({'success': True}), 200

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    """Delete admin user"""
    user = User.query.get_or_404(user_id)
    current_user = get_current_user()
    
    # Prevent deleting yourself
    if current_user and user.id == current_user.get('id'):
        return jsonify({'error': 'Cannot delete your own account'}), 400
    
    # Prevent deleting last admin
    if user.role == 'admin' and User.query.filter_by(role='admin').count() <= 1:
        return jsonify({'error': 'Cannot delete the last admin user'}), 400
    
    username = user.username
    db.session.delete(user)
    log_activity('DELETE', 'User', f'Deleted user: {username}', current_user.get('username') if current_user else 'system')
    db.session.commit()
    
    logger.info(f"User deleted: {username} by {current_user.get('username') if current_user else 'system'}")
    
    return jsonify({'success': True}), 200

@app.route('/api/users/<int:user_id>/smtp-password', methods=['PUT'])
@admin_required
def update_smtp_password(user_id):
    """Update user's SMTP password"""
    user = User.query.get_or_404(user_id)
    data = request.get_json() or {}
    
    smtp_password = data.get('smtp_password', '')
    if smtp_password:
        user.smtp_password = smtp_password
        log_activity('UPDATE', 'User', f'Updated SMTP password for user: {user.username}', get_current_user())
        db.session.commit()
    
    return jsonify({'success': True}), 200

# ══════════════════════════════════════════════════════════════════════════════
# TEMPORARY ASSIGNMENTS (Loaner Devices)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/temporary-assignments', methods=['GET'])
@token_required
def get_temporary_assignments():
    """Get all temporary assignments"""
    from models import TemporaryAssignment
    
    status = request.args.get('status', '').strip()
    
    q = TemporaryAssignment.query
    
    if status:
        q = q.filter_by(status=status)
    
    assignments = q.order_by(TemporaryAssignment.created_at.desc()).all()
    
    return jsonify({
        'assignments': [{
            'id': a.id,
            'employee_id': a.employee_id,
            'employee_name': a.employee_name,
            'original_asset_id': a.original_asset_id,
            'original_asset_name': a.original_asset_name,
            'original_asset_serial': a.original_asset_serial,
            'temp_asset_id': a.temp_asset_id,
            'temp_asset_name': a.temp_asset_name,
            'temp_asset_serial': a.temp_asset_serial,
            'reason': a.reason,
            'start_date': a.start_date.isoformat() if a.start_date else None,
            'expected_return_date': a.expected_return_date.isoformat() if a.expected_return_date else None,
            'actual_return_date': a.actual_return_date.isoformat() if a.actual_return_date else None,
            'status': a.status,
            'created_at': a.created_at.isoformat() if a.created_at else None,
        } for a in assignments]
    }), 200

@app.route('/api/temporary-assignments', methods=['POST'])
@token_required
def create_temporary_assignment():
    """Create new temporary assignment"""
    from models import TemporaryAssignment
    
    data = request.get_json() or {}
    current_user = get_current_user()
    current_username = current_user.get('username') if current_user else 'system'
    
    # Validate required fields
    required = ['employee_id', 'employee_name', 'original_asset_id', 'temp_asset_id', 'reason']
    for field in required:
        if not data.get(field):
            return jsonify({'error': f'{field} is required'}), 400
    
    # Get asset details
    original_asset = Asset.query.get(data['original_asset_id'])
    temp_asset = Asset.query.get(data['temp_asset_id'])
    
    if not original_asset:
        return jsonify({'error': 'Original asset not found'}), 404
    if not temp_asset:
        return jsonify({'error': 'Temporary asset not found'}), 404
    
    # Check if temp asset is available
    if temp_asset.status != 'Available':
        return jsonify({'error': 'Temporary asset is not available'}), 400
    
    # Create assignment
    assignment = TemporaryAssignment(
        employee_id=data['employee_id'],
        employee_name=data['employee_name'],
        original_asset_id=original_asset.id,
        original_asset_name=original_asset.asset_name,
        original_asset_serial=original_asset.serial_number,
        temp_asset_id=temp_asset.id,
        temp_asset_name=temp_asset.asset_name,
        temp_asset_serial=temp_asset.serial_number,
        reason=data['reason'],
        start_date=parse_date(data.get('start_date')) or date.today(),
        expected_return_date=parse_date(data.get('expected_return_date')),
        status='Active'
    )
    db.session.add(assignment)
    
    # Update asset statuses
    original_asset.status = 'Maintenance'
    temp_asset.status = 'Assigned'
    temp_asset.emp_id = data['employee_id']
    temp_asset.employee_name = data['employee_name']
    
    # Create audit logs
    AuditService.log(
        action_type='TEMP_ASSIGNMENT_CREATED',
        module='TemporaryAssignment',
        asset_id=temp_asset.id,
        asset_name=temp_asset.asset_name,
        asset_serial=temp_asset.serial_number,
        category=temp_asset.category,
        employee_id=data['employee_id'],
        employee_name=data['employee_name'],
        old_value=f'Original: {original_asset.asset_name}',
        new_value=f'Temp: {temp_asset.asset_name}',
        performed_by=current_username,
        remarks=f"Reason: {data['reason']}"
    )
    
    # Create lifecycle events
    LifecycleService.record_event(
        asset_id=original_asset.id,
        event_type='MAINTENANCE_STARTED',
        from_status='Assigned',
        to_status='Maintenance',
        reason=data['reason'],
        performed_by=current_username
    )
    
    LifecycleService.record_event(
        asset_id=temp_asset.id,
        event_type='TEMP_ASSIGNED',
        to_employee_id=data['employee_id'],
        to_employee=data['employee_name'],
        from_status='Available',
        to_status='Assigned',
        reason=f"Temporary replacement for {original_asset.asset_name}",
        performed_by=current_username
    )
    
    log_activity('CREATE', 'TemporaryAssignment', 
                f'Created temporary assignment: {temp_asset.asset_name} for {data["employee_name"]}', 
                current_username)
    db.session.commit()
    
    return jsonify({'success': True, 'assignment': assignment.id}), 201

@app.route('/api/temporary-assignments/<int:assignment_id>/complete', methods=['POST'])
@token_required
def complete_temporary_assignment(assignment_id):
    """Complete a temporary assignment and return assets to normal"""
    from models import TemporaryAssignment
    
    assignment = TemporaryAssignment.query.get_or_404(assignment_id)
    current_user = get_current_user()
    current_username = current_user.get('username') if current_user else 'system'
    
    if assignment.status != 'Active':
        return jsonify({'error': 'Assignment is not active'}), 400
    
    # Get assets
    original_asset = Asset.query.get(assignment.original_asset_id)
    temp_asset = Asset.query.get(assignment.temp_asset_id)
    
    # Update assignment
    assignment.status = 'Completed'
    assignment.actual_return_date = date.today()
    
    # Update asset statuses
    if original_asset:
        original_asset.status = 'Assigned'  # Back to assigned to employee
    
    if temp_asset:
        temp_asset.status = 'Available'  # Return to inventory
        temp_asset.emp_id = ''
        temp_asset.employee_name = ''
    
    # Create audit logs
    AuditService.log(
        action_type='TEMP_ASSIGNMENT_COMPLETED',
        module='TemporaryAssignment',
        asset_id=temp_asset.id if temp_asset else None,
        asset_name=temp_asset.asset_name if temp_asset else '',
        asset_serial=temp_asset.serial_number if temp_asset else '',
        employee_id=assignment.employee_id,
        employee_name=assignment.employee_name,
        performed_by=current_username,
        remarks=f"Completed temporary assignment. Original asset restored."
    )
    
    # Create lifecycle events
    if original_asset:
        LifecycleService.record_event(
            asset_id=original_asset.id,
            event_type='MAINTENANCE_COMPLETED',
            from_status='Maintenance',
            to_status='Assigned',
            performed_by=current_username
        )
    
    if temp_asset:
        LifecycleService.record_event(
            asset_id=temp_asset.id,
            event_type='RETURNED',
            from_employee_id=assignment.employee_id,
            from_employee=assignment.employee_name,
            from_status='Assigned',
            to_status='Available',
            performed_by=current_username
        )
    
    log_activity('UPDATE', 'TemporaryAssignment', 
                f'Completed temporary assignment for {assignment.employee_name}', 
                current_username)
    db.session.commit()
    
    return jsonify({'success': True}), 200

@app.route('/api/temporary-assignments/<int:assignment_id>', methods=['DELETE'])
@token_required
def delete_temporary_assignment(assignment_id):
    """Delete a temporary assignment"""
    from models import TemporaryAssignment
    
    assignment = TemporaryAssignment.query.get_or_404(assignment_id)
    current_user = get_current_user()
    current_username = current_user.get('username') if current_user else 'system'
    
    employee_name = assignment.employee_name
    db.session.delete(assignment)
    log_activity('DELETE', 'TemporaryAssignment', 
                f'Deleted temporary assignment for {employee_name}', 
                current_username)
    db.session.commit()
    
    return jsonify({'success': True}), 200

@app.route('/api/assets/<int:asset_id>/details', methods=['GET'])
def get_asset_details(asset_id):
    """Get detailed asset information"""
    asset = Asset.query.get_or_404(asset_id)
    return jsonify({'asset': asset.to_dict()}), 200

@app.route('/api/assets/by-employee/<emp_id>', methods=['GET'])
@token_required
def get_assets_by_employee(emp_id):
    """Get all assets assigned to an employee"""
    assets = Asset.query.filter_by(emp_id=emp_id).all()
    
    employee_name = ''
    if assets:
        employee_name = assets[0].employee_name
    
    return jsonify({
        'assets': [a.to_dict() for a in assets],
        'employee_name': employee_name
    }), 200

@app.route('/api/assets/<int:asset_id>/history', methods=['GET'])
def get_asset_history(asset_id):
    """Get complete lifecycle history for an asset"""
    from models import AssetLifecycle, AuditLog, TemporaryAssignment
    
    asset = Asset.query.get_or_404(asset_id)
    
    # Get lifecycle events
    lifecycle_events = AssetLifecycle.query.filter_by(asset_id=asset_id)\
        .order_by(AssetLifecycle.created_at.desc()).all()
    
    # Get audit logs for this asset
    audit_logs = AuditLog.query.filter_by(asset_id=asset_id)\
        .order_by(AuditLog.timestamp.desc()).limit(50).all()
    
    # Get temporary assignments where this was original or temp asset
    temp_assignments = TemporaryAssignment.query.filter(
        or_(
            TemporaryAssignment.original_asset_id == asset_id,
            TemporaryAssignment.temp_asset_id == asset_id
        )
    ).order_by(TemporaryAssignment.created_at.desc()).all()
    
    # Combine and sort all events by date
    all_events = []
    
    # Add lifecycle events
    for event in lifecycle_events:
        all_events.append({
            'type': 'lifecycle',
            'event_type': event.event_type,
            'date': event.created_at.isoformat() if event.created_at else '',
            'from_employee': event.from_employee,
            'to_employee': event.to_employee,
            'from_status': event.from_status,
            'to_status': event.to_status,
            'reason': event.reason,
            'performed_by': event.performed_by,
            'remarks': event.remarks,
        })
    
    # Add key audit logs (assignments, returns, status changes)
    for log in audit_logs:
        if log.action_type in ['ASSET_CREATED', 'ASSET_ASSIGNED', 'ASSET_RETURNED', 
                                'ASSET_REASSIGNED', 'STATUS_CHANGED', 'TEMP_ASSIGNMENT_CREATED',
                                'TEMP_ASSIGNMENT_COMPLETED', 'ASSET_REPLACED']:
            all_events.append({
                'type': 'audit',
                'action_type': log.action_type,
                'date': log.timestamp.isoformat() if log.timestamp else '',
                'employee_name': log.employee_name,
                'field_name': log.field_name,
                'old_value': log.old_value,
                'new_value': log.new_value,
                'performed_by': log.performed_by,
                'remarks': log.remarks,
            })
    
    # Add temporary assignments
    for assignment in temp_assignments:
        if assignment.original_asset_id == asset_id:
            all_events.append({
                'type': 'temp_assignment',
                'sub_type': 'original',
                'date': assignment.created_at.isoformat() if assignment.created_at else '',
                'employee_name': assignment.employee_name,
                'temp_asset_name': assignment.temp_asset_name,
                'reason': assignment.reason,
                'status': assignment.status,
                'start_date': assignment.start_date.isoformat() if assignment.start_date else '',
                'expected_return': assignment.expected_return_date.isoformat() if assignment.expected_return_date else '',
                'actual_return': assignment.actual_return_date.isoformat() if assignment.actual_return_date else '',
            })
        else:
            all_events.append({
                'type': 'temp_assignment',
                'sub_type': 'temporary',
                'date': assignment.created_at.isoformat() if assignment.created_at else '',
                'employee_name': assignment.employee_name,
                'original_asset_name': assignment.original_asset_name,
                'reason': assignment.reason,
                'status': assignment.status,
                'start_date': assignment.start_date.isoformat() if assignment.start_date else '',
                'expected_return': assignment.expected_return_date.isoformat() if assignment.expected_return_date else '',
                'actual_return': assignment.actual_return_date.isoformat() if assignment.actual_return_date else '',
            })
    
    # Sort all events by date (newest first)
    all_events.sort(key=lambda x: x['date'], reverse=True)
    
    return jsonify({
        'asset': asset.to_dict(),
        'history': all_events,
        'total_events': len(all_events),
        'lifecycle_events_count': len(lifecycle_events),
        'audit_logs_count': len(audit_logs),
        'temp_assignments_count': len(temp_assignments),
    }), 200

# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/dashboard/stats', methods=['GET'])
def dashboard_stats():
    total       = Asset.query.count()
    assigned    = Asset.query.filter_by(status='Assigned').count()
    available   = Asset.query.filter_by(status='Available').count()
    maintenance = Asset.query.filter_by(status='Maintenance').count()

    today = date.today()
    soon  = today + timedelta(days=90)
    expiring = Asset.query.filter(
        Asset.warranty_date != None,
        Asset.warranty_date <= soon,
        Asset.warranty_date >= today
    ).count()

    # Category breakdown (only assigned assets)
    from sqlalchemy import func
    cat_rows = db.session.query(Asset.category, func.count(Asset.id))\
                         .filter(Asset.status == 'Assigned')\
                         .group_by(Asset.category).all()
    categories = [{'name': r[0] or 'Unknown', 'count': r[1]} for r in cat_rows]

    # Laptop status breakdown
    laptop_total = Asset.query.filter_by(category='Laptop').count()
    laptop_available = Asset.query.filter_by(category='Laptop', status='Available').count()
    laptop_assigned = Asset.query.filter_by(category='Laptop', status='Assigned').count()
    laptop_maintenance = Asset.query.filter_by(category='Laptop', status='Maintenance').count()
    laptop_retired = Asset.query.filter_by(category='Laptop', status='Retired').count()

    return jsonify({
        'totalAssets':       total,
        'assignedAssets':    assigned,
        'availableAssets':   available,
        'maintenanceAssets': maintenance,
        'expiringWarranties': expiring,
        'categories':        categories,
        'laptopStats': {
            'total': laptop_total,
            'available': laptop_available,
            'assigned': laptop_assigned,
            'maintenance': laptop_maintenance,
            'retired': laptop_retired
        }
    }), 200

@app.route('/api/dashboard/activity', methods=['GET'])
def dashboard_activity():
    logs = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).limit(10).all()
    return jsonify({'logs': [l.to_dict() for l in logs]}), 200

@app.route('/api/dashboard/lifecycle-stats', methods=['GET'])
@non_viewer_required
def lifecycle_stats():
    from sqlalchemy import func
    today = date.today()
    month_start = today.replace(day=1)
    
    try:
        from models import AssetLifecycleEvent
        active_temp = AssetLifecycleEvent.query.filter_by(event_type='TEMP_ASSIGNED').count()
        under_repair = Asset.query.filter_by(status='Maintenance').count()
        replaced = AssetLifecycleEvent.query.filter(
            AssetLifecycleEvent.event_type == 'REPLACED',
            AssetLifecycleEvent.created_at >= month_start
        ).count()
        total_events = AssetLifecycleEvent.query.count()
    except Exception:
        active_temp = 0
        under_repair = Asset.query.filter_by(status='Maintenance').count()
        replaced = 0
        total_events = 0

    return jsonify({
        'stats': {
            'active_temp_assignments': active_temp,
            'assets_under_repair': under_repair,
            'assets_replaced_this_month': replaced,
            'total_lifecycle_events': total_events,
        }
    }), 200

# ══════════════════════════════════════════════════════════════════════════════
# ASSET ROUTES  – full CRUD
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/assets', methods=['GET'])
@token_required
def get_assets():
    search   = request.args.get('search', '').strip()
    location = request.args.get('location', '').strip()
    category = request.args.get('category', '').strip()
    status   = request.args.get('status', '').strip()
    sort     = request.args.get('sort', 'id_desc').strip()
    page     = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    q = Asset.query

    if search:
        q = q.filter(or_(
            Asset.asset_name.ilike(f'%{search}%'),
            Asset.serial_number.ilike(f'%{search}%'),
            Asset.emp_id.ilike(f'%{search}%'),
            Asset.employee_name.ilike(f'%{search}%'),
            Asset.model_name.ilike(f'%{search}%'),
        ))
    if location:
        q = q.filter(Asset.location.ilike(f'%{location}%'))
    if category:
        q = q.filter(Asset.category.ilike(f'%{category}%'))
    if status:
        q = q.filter_by(status=status)

    # Apply sorting - "Last Added" now sorts by updated_at DESC to show recently updated items first
    sort_column = Asset.updated_at.desc()  # Default: most recently updated first
    
    if sort == 'id_asc':
        sort_column = Asset.id.asc()
    elif sort == 'id_desc':
        sort_column = Asset.updated_at.desc()  # Changed from created_at to updated_at
    elif sort == 'emp_asc':
        sort_column = Asset.emp_id.asc()
    elif sort == 'emp_desc':
        sort_column = Asset.emp_id.desc()
    elif sort == 'name_asc':
        sort_column = Asset.asset_name.asc()
    elif sort == 'name_desc':
        sort_column = Asset.asset_name.desc()
    
    total  = q.count()
    assets = q.order_by(sort_column).offset((page-1)*per_page).limit(per_page).all()

    return jsonify({
        'assets': [a.to_dict() for a in assets],
        'total':  total,
        'page':   page,
        'pages':  (total + per_page - 1) // per_page,
    }), 200

@app.route('/api/assets/<int:asset_id>', methods=['GET'])
@token_required
def get_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    return jsonify(asset.to_dict()), 200

@app.route('/api/assets', methods=['POST'])
@non_viewer_required
def create_asset():
    data = request.get_json() or {}
    current_user = get_current_user()

    # Validate required fields
    if not data.get('asset_name') or not data.get('serial_number'):
        return jsonify({'error': 'asset_name and serial_number are required'}), 400

    # Check duplicate serial number
    if Asset.query.filter_by(serial_number=data['serial_number'].strip()).first():
        return jsonify({'error': 'Serial number already exists'}), 409

    # Handle employee_email field (accept both employee_email and email)
    employee_email = data.get('employee_email') or data.get('email', '')

    asset = Asset(
        emp_id          = data.get('emp_id', ''),
        employee_name   = data.get('employee_name', ''),
        mobile_number   = data.get('mobile_number', ''),
        employee_email  = employee_email,
        asset_name      = data['asset_name'].strip(),
        category        = data.get('category', ''),
        serial_number   = data['serial_number'].strip(),
        model_name      = data.get('model_name', ''),
        os              = data.get('os', ''),
        version         = data.get('version', ''),
        ram             = data.get('ram', ''),
        location        = data.get('location', ''),
        invoice_number  = data.get('invoice_number', ''),
        invoice_date    = parse_date(data.get('invoice_date')),
        warranty_date   = parse_date(data.get('warranty_date')),
        charger_serial  = data.get('charger_serial', ''),
        old_user        = data.get('old_user', ''),
        date            = parse_date(data.get('date')) or date.today(),
        old_device      = data.get('old_device', ''),
        comments        = data.get('comments', ''),
        status          = data.get('status', 'Available'),
        # Legacy inventory fields
        purchase_price       = data.get('purchase_price'),
        quantity             = data.get('quantity', 1),
        configuration        = data.get('configuration', ''),
        laptop_bag_serial    = data.get('laptop_bag_serial', ''),
        hard_disk_serial     = data.get('hard_disk_serial', ''),
        hard_disk_capacity   = data.get('hard_disk_capacity', ''),
        ups_serial           = data.get('ups_serial', ''),
        ups_capacity         = data.get('ups_capacity', ''),
        printer_type         = data.get('printer_type', ''),
        printer_model        = data.get('printer_model', ''),
        mobile_imei          = data.get('mobile_imei', ''),
        mobile_number_sim    = data.get('mobile_number_sim', ''),
        testing_status       = data.get('testing_status', ''),
        # New dynamic category-specific fields
        brand_name           = data.get('brand_name', ''),
        processor            = data.get('processor', ''),
        storage_type         = data.get('storage_type', ''),
        storage_capacity     = data.get('storage_capacity', ''),
        graphics_card        = data.get('graphics_card', ''),
        os_version           = data.get('os_version', ''),
        screen_size          = data.get('screen_size', ''),
        imei_1               = data.get('imei_1', ''),
        imei_2               = data.get('imei_2', ''),
        color_or_mono        = data.get('color_or_mono', ''),
        network_enabled      = data.get('network_enabled', ''),
        resolution           = data.get('resolution', ''),
        refresh_rate         = data.get('refresh_rate', ''),
        cpu_count            = data.get('cpu_count'),
        raid_config          = data.get('raid_config', ''),
        ip_address           = data.get('ip_address', ''),
        rack_location        = data.get('rack_location', ''),
        interface_type       = data.get('interface_type', ''),
        capacity_va          = data.get('capacity_va', ''),
        battery_type         = data.get('battery_type', ''),
        backup_time          = data.get('backup_time', ''),
        connection_type      = data.get('connection_type', ''),
        noise_cancellation   = data.get('noise_cancellation', ''),
        size_compatibility   = data.get('size_compatibility', ''),
        color                = data.get('color', ''),
        warranty_period      = data.get('warranty_period', ''),
        purchase_vendor      = data.get('purchase_vendor', ''),
        purchase_date        = parse_date(data.get('purchase_date')),
        warranty_start_date  = parse_date(data.get('warranty_start_date')),
        warranty_end_date    = parse_date(data.get('warranty_end_date')),
        assigned_employee    = data.get('assigned_employee', ''),
        custom_description   = data.get('custom_description', ''),
        remarks              = data.get('remarks', ''),
    )
    db.session.add(asset)
    
    current_username = current_user.get('username') if current_user else 'system'
    log_activity('CREATE', 'Asset', f'Added asset: {asset.asset_name} [{asset.serial_number}]', current_username)
    db.session.flush()  # Flush to get asset.id before commit

    # Create comprehensive audit log
    try:
        audit_log = AuditService.log_asset_created(asset, current_username)
        logger.info(f"Asset created: {asset.asset_name} (ID: {asset.id}) by {current_username}")
    except Exception as e:
        logger.error(f"Error creating audit log for asset {asset.id}: {e}")
    
    # Create lifecycle event
    try:
        lifecycle_event = LifecycleService.record_event(
            asset_id=asset.id,
            event_type='PROCURED',
            to_status=asset.status,
            reason='New asset added to inventory',
            performed_by=current_username
        )
        logger.info(f"Lifecycle event created for asset {asset.id}")
    except Exception as e:
        logger.error(f"Error creating lifecycle event for asset {asset.id}: {e}")
    
    db.session.commit()  # Commit everything together
    logger.info(f"Asset {asset.id} successfully committed to database")

    return jsonify({'success': True, 'asset': asset.to_dict()}), 201

@app.route('/api/assets/<int:asset_id>', methods=['PUT'])
@non_viewer_required
def update_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    data  = request.get_json() or {}
    current_user = get_current_user()
    current_username = current_user.get('username') if current_user else 'system'

    # Track changes for audit log
    changed_fields = {}
    old_status = asset.status

    # Check serial number uniqueness if changed
    new_serial = data.get('serial_number', asset.serial_number).strip()
    if new_serial != asset.serial_number:
        if Asset.query.filter_by(serial_number=new_serial).first():
            return jsonify({'error': 'Serial number already exists'}), 409
        changed_fields['serial_number'] = (asset.serial_number, new_serial)

    # Handle employee_email field (accept both employee_email and email)
    if 'employee_email' in data or 'email' in data:
        new_email = data.get('employee_email') or data.get('email', '')
        if new_email != asset.employee_email:
            changed_fields['employee_email'] = (asset.employee_email, new_email)
        asset.employee_email = new_email

    # Track all field changes
    fields_to_track = {
        'emp_id': 'emp_id',
        'employee_name': 'employee_name',
        'mobile_number': 'mobile_number',
        'asset_name': 'asset_name',
        'category': 'category',
        'model_name': 'model_name',
        'os': 'os',
        'version': 'version',
        'ram': 'ram',
        'location': 'location',
        'invoice_number': 'invoice_number',
        'charger_serial': 'charger_serial',
        'old_user': 'old_user',
        'old_device': 'old_device',
        'comments': 'comments',
        'status': 'status',
    }
    
    for field_key, field_name in fields_to_track.items():
        if field_key in data:
            old_val = getattr(asset, field_name)
            new_val = data[field_key]
            if field_key == 'asset_name':
                new_val = new_val.strip()
            if str(old_val) != str(new_val):
                changed_fields[field_name] = (old_val, new_val)

    asset.emp_id         = data.get('emp_id',         asset.emp_id)
    asset.employee_name  = data.get('employee_name',  asset.employee_name)
    asset.mobile_number  = data.get('mobile_number',  asset.mobile_number)
    asset.asset_name     = data.get('asset_name',     asset.asset_name).strip()
    asset.category       = data.get('category',       asset.category)
    asset.serial_number  = new_serial
    asset.model_name     = data.get('model_name',     asset.model_name)
    asset.os             = data.get('os',             asset.os)
    asset.version        = data.get('version',        asset.version)
    asset.ram            = data.get('ram',            asset.ram)
    asset.location       = data.get('location',       asset.location)
    asset.invoice_number = data.get('invoice_number', asset.invoice_number)
    asset.invoice_date   = parse_date(data.get('invoice_date')) or asset.invoice_date
    asset.warranty_date  = parse_date(data.get('warranty_date')) or asset.warranty_date
    asset.charger_serial = data.get('charger_serial', asset.charger_serial)
    asset.old_user       = data.get('old_user',       asset.old_user)
    asset.date           = parse_date(data.get('date')) or asset.date
    asset.old_device     = data.get('old_device',     asset.old_device)
    asset.comments       = data.get('comments',       asset.comments)
    asset.status         = data.get('status',         asset.status)
    asset.updated_at     = datetime.utcnow()
    
    # Legacy inventory fields
    if 'purchase_price' in data:
        asset.purchase_price = safe_float(data.get('purchase_price'))
    if 'quantity' in data:
        asset.quantity = safe_int(data.get('quantity'), 1)
    if 'configuration' in data:
        asset.configuration = data.get('configuration', '')
    if 'laptop_bag_serial' in data:
        asset.laptop_bag_serial = data.get('laptop_bag_serial', '')
    if 'hard_disk_serial' in data:
        asset.hard_disk_serial = data.get('hard_disk_serial', '')
    if 'hard_disk_capacity' in data:
        asset.hard_disk_capacity = data.get('hard_disk_capacity', '')
    if 'ups_serial' in data:
        asset.ups_serial = data.get('ups_serial', '')
    if 'ups_capacity' in data:
        asset.ups_capacity = data.get('ups_capacity', '')
    if 'printer_type' in data:
        asset.printer_type = data.get('printer_type', '')
    if 'printer_model' in data:
        asset.printer_model = data.get('printer_model', '')
    if 'mobile_imei' in data:
        asset.mobile_imei = data.get('mobile_imei', '')
    if 'mobile_number_sim' in data:
        asset.mobile_number_sim = data.get('mobile_number_sim', '')
    if 'testing_status' in data:
        asset.testing_status = data.get('testing_status', '')
    
    # New dynamic category-specific fields
    if 'brand_name' in data:
        asset.brand_name = data.get('brand_name', '')
    if 'processor' in data:
        asset.processor = data.get('processor', '')
    if 'storage_type' in data:
        asset.storage_type = data.get('storage_type', '')
    if 'storage_capacity' in data:
        asset.storage_capacity = data.get('storage_capacity', '')
    if 'graphics_card' in data:
        asset.graphics_card = data.get('graphics_card', '')
    if 'os_version' in data:
        asset.os_version = data.get('os_version', '')
    if 'screen_size' in data:
        asset.screen_size = data.get('screen_size', '')
    if 'imei_1' in data:
        asset.imei_1 = data.get('imei_1', '')
    if 'imei_2' in data:
        asset.imei_2 = data.get('imei_2', '')
    if 'mobile_number' in data:
        asset.mobile_number = data.get('mobile_number', '')
    if 'color_or_mono' in data:
        asset.color_or_mono = data.get('color_or_mono', '')
    if 'network_enabled' in data:
        asset.network_enabled = data.get('network_enabled', '')
    if 'resolution' in data:
        asset.resolution = data.get('resolution', '')
    if 'refresh_rate' in data:
        asset.refresh_rate = data.get('refresh_rate', '')
    if 'cpu_count' in data:
        asset.cpu_count = safe_int(data.get('cpu_count'))
    if 'raid_config' in data:
        asset.raid_config = data.get('raid_config', '')
    if 'ip_address' in data:
        asset.ip_address = data.get('ip_address', '')
    if 'rack_location' in data:
        asset.rack_location = data.get('rack_location', '')
    if 'interface_type' in data:
        asset.interface_type = data.get('interface_type', '')
    if 'capacity_va' in data:
        asset.capacity_va = data.get('capacity_va', '')
    if 'battery_type' in data:
        asset.battery_type = data.get('battery_type', '')
    if 'backup_time' in data:
        asset.backup_time = data.get('backup_time', '')
    if 'connection_type' in data:
        asset.connection_type = data.get('connection_type', '')
    if 'noise_cancellation' in data:
        asset.noise_cancellation = data.get('noise_cancellation', '')
    if 'size_compatibility' in data:
        asset.size_compatibility = data.get('size_compatibility', '')
    if 'color' in data:
        asset.color = data.get('color', '')
    if 'warranty_period' in data:
        asset.warranty_period = data.get('warranty_period', '')
    if 'purchase_vendor' in data:
        asset.purchase_vendor = data.get('purchase_vendor', '')
    if 'purchase_date' in data:
        asset.purchase_date = parse_date(data.get('purchase_date'))
    if 'warranty_start_date' in data:
        asset.warranty_start_date = parse_date(data.get('warranty_start_date'))
    if 'warranty_end_date' in data:
        asset.warranty_end_date = parse_date(data.get('warranty_end_date'))
    if 'assigned_employee' in data:
        asset.assigned_employee = data.get('assigned_employee', '')
    if 'custom_description' in data:
        asset.custom_description = data.get('custom_description', '')
    if 'remarks' in data:
        asset.remarks = data.get('remarks', '')

    log_activity('UPDATE', 'Asset', f'Updated asset: {asset.asset_name} [{asset.serial_number}]', current_username)

    # Create comprehensive audit logs for field changes
    if changed_fields:
        AuditService.log_asset_updated(asset, changed_fields, current_username)
    
    # If status changed, create additional status change log and lifecycle event
    if 'status' in changed_fields:
        new_status = changed_fields['status'][1]
        AuditService.log_status_change(asset, old_status, new_status, current_username)
        
        LifecycleService.record_event(
            asset_id=asset.id,
            event_type='STATUS_CHANGED',
            from_status=old_status,
            to_status=new_status,
            performed_by=current_username
        )
    
    # If employee changed (assignment/return)
    if 'emp_id' in changed_fields or 'employee_name' in changed_fields:
        old_emp = changed_fields.get('emp_id', (asset.emp_id, asset.emp_id))[0]
        new_emp = asset.emp_id
        
        if old_emp and not new_emp:  # Asset returned
            AuditService.log_asset_returned(
                asset, asset.employee_name or '', old_emp, current_username,
                new_status=asset.status
            )
            LifecycleService.record_event(
                asset_id=asset.id,
                event_type='RETURNED',
                from_employee_id=old_emp,
                from_employee=changed_fields.get('employee_name', (asset.employee_name, ''))[0],
                from_status='Assigned',
                to_status=asset.status,
                performed_by=current_username
            )
        elif new_emp and not old_emp:  # Asset assigned
            AuditService.log_asset_assigned(
                asset, asset.employee_name, new_emp, current_username,
                old_status=old_status
            )
            LifecycleService.record_event(
                asset_id=asset.id,
                event_type='ASSIGNED',
                to_employee_id=new_emp,
                to_employee=asset.employee_name,
                from_status=old_status,
                to_status='Assigned',
                performed_by=current_username
            )
        elif new_emp and old_emp and new_emp != old_emp:  # Reassignment
            AuditService.log(
                action_type='ASSET_REASSIGNED',
                module='Asset',
                asset_id=asset.id,
                asset_name=asset.asset_name,
                asset_serial=asset.serial_number,
                category=asset.category,
                employee_id=new_emp,
                employee_name=asset.employee_name,
                old_value=old_emp,
                new_value=new_emp,
                performed_by=current_username,
                remarks=f"Reassigned from {old_emp} to {new_emp}"
            )
            LifecycleService.record_event(
                asset_id=asset.id,
                event_type='REASSIGNED',
                from_employee_id=old_emp,
                to_employee_id=new_emp,
                to_employee=asset.employee_name,
                from_status=old_status,
                to_status=asset.status,
                performed_by=current_username
            )

    # Commit all changes (asset updates + activity log + audit logs + lifecycle events)
    db.session.commit()

    return jsonify({'success': True, 'asset': asset.to_dict()}), 200

@app.route('/api/assets/<int:asset_id>', methods=['DELETE'])
@token_required
def delete_asset(asset_id):
    from models import AssetLifecycle, AssetReplacement, TemporaryAssignment, ExitAssetCollection, OnboardingAssetAssignment
    
    asset = Asset.query.get_or_404(asset_id)
    current_user = get_current_user()
    name = asset.asset_name
    serial = asset.serial_number
    category = asset.category
    
    # Get username from current_user dict
    username = current_user.get('username') if current_user else 'system'
    
    # Create audit log before deletion
    AuditService.log_asset_deleted(asset, username)
    
    # Delete ALL related records first to avoid foreign key constraints
    # 1. Delete lifecycle events
    AssetLifecycle.query.filter_by(asset_id=asset_id).delete()
    
    # 2. Delete asset replacements where this asset is involved (old or new)
    AssetReplacement.query.filter(
        (AssetReplacement.old_asset_id == asset_id) | 
        (AssetReplacement.new_asset_id == asset_id)
    ).delete(synchronize_session=False)
    
    # 3. Delete temporary assignments where this asset is involved (original or temp)
    TemporaryAssignment.query.filter(
        (TemporaryAssignment.original_asset_id == asset_id) |
        (TemporaryAssignment.temp_asset_id == asset_id)
    ).delete(synchronize_session=False)
    
    # 4. Delete exit asset collection records
    ExitAssetCollection.query.filter_by(asset_id=asset_id).delete()
    
    # 5. Delete onboarding asset assignments
    OnboardingAssetAssignment.query.filter_by(asset_id=asset_id).delete()
    
    # Delete the asset
    db.session.delete(asset)
    log_activity('DELETE', 'Asset', f'Deleted asset: {name} [{serial}]', username)
    db.session.commit()
    
    logger.info(f"Asset deleted: {name} [{serial}] (ID: {asset_id}) by {username}")
    
    return jsonify({'success': True, 'message': f'Asset "{name}" deleted'}), 200

# ══════════════════════════════════════════════════════════════════════════════
# ASSET IMPORT/EXPORT
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/assets/template', methods=['GET'])
def download_asset_template():
    """Download Excel template for bulk asset import"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assets"
        
        # Define headers - matching user's Excel format exactly
        headers = [
            'Sl no.', 'EMP ID', 'EMPLOYEE NAME', 'MOBILE NUMBER', 'Asset NAME',
            'CATEGORY', 'SERIAL NUMBER', 'MODEL NAME', 'OS', 'Version', 'Ram',
            'LOCATION', 'INVOICE NUMBER', 'INVOICE DATE', 'WARRANTY DATE',
            'Charger Serial Number', 'Old User', 'Date', 'Old Device', 'Comments'
        ]
        
        # Style header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Add sample data rows matching user's format
        sample_data = [
            ['1', 'EMP001', 'John Doe', '1234567890', 'Dell Laptop XPS 15',
             'Laptop', 'SN-DELL-001', 'XPS 15 9500', 'Windows', '11', '16GB',
             'HQ Office', 'INV-001', '2024-01-15', '2027-01-15',
             'CHG-001', '', '2024-01-15', '', 'Primary work laptop'],
            ['2', '', '', '', 'HP Monitor 27"',
             'Monitor', 'SN-MON-002', 'HP E27', '', '', '',
             'HQ Office', 'INV-002', '2024-02-20', '2027-02-20',
             '', '', '2024-02-20', '', 'External display']
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='asset_import_template.xlsx'
        )
    except ImportError:
        return jsonify({
            'error': 'openpyxl library not installed. Please install it: pip install openpyxl'
        }), 500
    except Exception as e:
        logger.error(f"Error generating template: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/assets/import', methods=['POST'])
@token_required
def import_assets():
    """Bulk import assets from Excel file"""
    try:
        import openpyxl
        from datetime import datetime
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload .xlsx or .xls file'}), 400
        
        current_user = get_current_user()
        current_username = current_user.get('username') if current_user else 'system'
        
        # Read Excel file
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        imported_count = 0
        error_count = 0
        error_details = []
        imported_ids = []  # Track imported asset IDs
        
        # Process each row (skip header)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Create dict from row data
                data = dict(zip(headers, row))
                
                # Skip empty rows
                if not data.get('Asset NAME') and not data.get('SERIAL NUMBER'):
                    continue
                
                # Validate required fields
                if not data.get('Asset NAME'):
                    error_details.append(f"Row {row_num}: Missing Asset NAME")
                    error_count += 1
                    continue
                
                if not data.get('SERIAL NUMBER'):
                    error_details.append(f"Row {row_num}: Missing SERIAL NUMBER")
                    error_count += 1
                    continue
                
                # Check for duplicate serial number
                if Asset.query.filter_by(serial_number=str(data['SERIAL NUMBER']).strip()).first():
                    error_details.append(f"Row {row_num}: Serial number '{data['SERIAL NUMBER']}' already exists")
                    error_count += 1
                    continue
                
                # Parse dates
                invoice_date = None
                if data.get('INVOICE DATE'):
                    if isinstance(data['INVOICE DATE'], datetime):
                        invoice_date = data['INVOICE DATE'].date()
                    else:
                        try:
                            invoice_date = parse_date(str(data['INVOICE DATE']))
                        except:
                            pass
                
                warranty_date = None
                if data.get('WARRANTY DATE'):
                    if isinstance(data['WARRANTY DATE'], datetime):
                        warranty_date = data['WARRANTY DATE'].date()
                    else:
                        try:
                            warranty_date = parse_date(str(data['WARRANTY DATE']))
                        except:
                            pass
                
                assignment_date = None
                if data.get('Date'):
                    if isinstance(data['Date'], datetime):
                        assignment_date = data['Date'].date()
                    else:
                        try:
                            assignment_date = parse_date(str(data['Date']))
                        except:
                            pass
                
                # Determine status - default to Assigned if employee info exists, otherwise Available
                emp_id = str(data.get('EMP ID', '')).strip() if data.get('EMP ID') else ''
                emp_name = str(data.get('EMPLOYEE NAME', '')).strip() if data.get('EMPLOYEE NAME') else ''
                
                # If either EMP ID or EMPLOYEE NAME is filled, mark as Assigned
                # Otherwise, mark as Available
                if emp_id or emp_name:
                    asset_status = 'Assigned'
                else:
                    asset_status = 'Available'
                
                # Create asset - mapping user's columns to database fields
                asset = Asset(
                    asset_name=str(data.get('Asset NAME', '')).strip(),
                    serial_number=str(data.get('SERIAL NUMBER', '')).strip(),
                    category=str(data.get('CATEGORY', '')).strip() if data.get('CATEGORY') else '',
                    model_name=str(data.get('MODEL NAME', '')).strip() if data.get('MODEL NAME') else '',
                    os=str(data.get('OS', '')).strip() if data.get('OS') else '',
                    version=str(data.get('Version', '')).strip() if data.get('Version') else '',
                    ram=str(data.get('Ram', '')).strip() if data.get('Ram') else '',
                    location=str(data.get('LOCATION', '')).strip() if data.get('LOCATION') else '',
                    invoice_number=str(data.get('INVOICE NUMBER', '')).strip() if data.get('INVOICE NUMBER') else '',
                    invoice_date=invoice_date,
                    warranty_date=warranty_date,
                    charger_serial=str(data.get('Charger Serial Number', '')).strip() if data.get('Charger Serial Number') else '',
                    old_user=str(data.get('Old User', '')).strip() if data.get('Old User') else '',
                    date=assignment_date or date.today(),
                    old_device=str(data.get('Old Device', '')).strip() if data.get('Old Device') else '',
                    comments=str(data.get('Comments', '')).strip() if data.get('Comments') else '',
                    emp_id=emp_id,
                    employee_name=emp_name,
                    mobile_number=str(data.get('MOBILE NUMBER', '')).strip() if data.get('MOBILE NUMBER') else '',
                    status=asset_status
                )
                
                
                
                db.session.add(asset)
                db.session.flush()  # Flush to get the asset ID
                imported_ids.append(asset.id)  # Track the ID
                imported_count += 1

               
               
                # Create audit log
                AuditService.log(
                    action_type='ASSET_IMPORTED',
                    module='Asset',
                    asset_id=None,
                    asset_name=asset.asset_name,
                    asset_serial=asset.serial_number,
                    category=asset.category,
                    performed_by=current_username,
                    remarks=f'Imported from Excel (Row {row_num})'
                )
                
            except Exception as e:
                error_details.append(f"Row {row_num}: {str(e)}")
                error_count += 1
                logger.error(f"Error importing row {row_num}: {e}")
        
        db.session.commit()
        
        message = f'Successfully imported {imported_count} assets'
        if error_count > 0:
            message += f', {error_count} rows had errors'
        
        return jsonify({
            'success': True,
            'message': message,
            'imported': imported_count,
            'errors': error_count,
            'error_details': error_details[:10],  # Limit to first 10 errors
            'imported_ids': imported_ids  # Return imported asset IDs
        }), 200
        
    except ImportError:
        return jsonify({
            'error': 'openpyxl library not installed. Please install it: pip install openpyxl'
        }), 500
    except Exception as e:
        logger.error(f"Error importing assets: {e}")
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# ══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/employees', methods=['GET'])
@admin_required
def get_employees():
    """Get all employees or search by query"""
    from models import Employee
    
    query = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    q = Employee.query
    
    if query:
        q = q.filter(or_(
            Employee.emp_id.ilike(f'%{query}%'),
            Employee.employee_name.ilike(f'%{query}%'),
            Employee.email.ilike(f'%{query}%'),
            Employee.mobile_number.ilike(f'%{query}%')
        ))
    
    total = q.count()
    employees = q.order_by(Employee.created_at.desc()).offset((page-1)*per_page).limit(per_page).all()
    
    return jsonify([{
        'emp_id': e.emp_id,
        'employee_name': e.employee_name,
        'email': e.email,
        'mobile_number': e.mobile_number,
        'location': e.location,
        'department': e.department,
        'designation': e.designation,
        'created_at': e.created_at.isoformat() if e.created_at else None,
    } for e in employees]), 200

@app.route('/api/employees/<emp_id>', methods=['GET'])
@admin_required
def get_employee(emp_id):
    """Get employee by emp_id"""
    from models import Employee
    
    employee = Employee.query.filter_by(emp_id=emp_id).first()
    if not employee:
        return jsonify({'found': False}), 404
    
    return jsonify({
        'found': True,
        'employee': {
            'emp_id': employee.emp_id,
            'employee_name': employee.employee_name,
            'email': employee.email,
            'mobile_number': employee.mobile_number,
            'location': employee.location,
            'department': employee.department,
            'designation': employee.designation,
            'status': getattr(employee, 'status', 'Active'),
        }
    }), 200

@app.route('/api/employees/<emp_id>/assets', methods=['GET'])
@admin_required
def get_employee_assets(emp_id):
    """Get all assets assigned to an employee"""
    from models import Asset
    
    assets = Asset.query.filter_by(emp_id=emp_id).all()
    
    return jsonify([{
        'id': a.id,
        'asset_name': a.asset_name,
        'category': a.category,
        'serial_number': a.serial_number,
        'model_name': a.model_name,
        'brand_name': a.brand_name,
        'status': a.status,
    } for a in assets]), 200

@app.route('/api/employees/<emp_id>/asset-history', methods=['GET'])
@admin_required
def get_employee_asset_history(emp_id):
    """Get complete asset history for an employee - every device they've ever used"""
    from models import Asset, AssetLifecycle, AuditLog, TemporaryAssignment, AssetReplacement, Employee
    from sqlalchemy import or_, and_
    
    # Get employee details
    employee = Employee.query.filter_by(emp_id=emp_id).first()
    if not employee:
        return jsonify({'error': 'Employee not found'}), 404
    
    # Get currently assigned assets
    current_assets = Asset.query.filter_by(emp_id=emp_id).all()
    
    # Collect all events for this employee from different sources
    events = []
    
    # 1. AssetLifecycle events where this employee is involved
    lifecycle_events = AssetLifecycle.query.filter(
        or_(
            AssetLifecycle.to_employee_id == emp_id,
            AssetLifecycle.from_employee_id == emp_id,
            AssetLifecycle.to_employee == employee.employee_name,
            AssetLifecycle.from_employee == employee.employee_name
        )
    ).order_by(AssetLifecycle.event_date.desc()).all()
    
    for event in lifecycle_events:
        asset = Asset.query.get(event.asset_id) if event.asset_id else None
        events.append({
            'type': 'lifecycle',
            'event_type': event.event_type,
            'date': event.event_date.isoformat() if event.event_date else None,
            'timestamp': event.event_date.isoformat() if event.event_date else None,
            'asset_id': event.asset_id,
            'asset_name': asset.asset_name if asset else None,
            'asset_serial': asset.serial_number if asset else None,
            'category': asset.category if asset else None,
            'brand_name': asset.brand_name if asset else None,
            'model_name': asset.model_name if asset else None,
            'to_employee': event.to_employee,
            'to_employee_id': event.to_employee_id,
            'from_employee': event.from_employee,
            'from_employee_id': event.from_employee_id,
            'from_status': event.from_status,
            'to_status': event.to_status,
            'reason': event.reason,
            'location': event.location,
            'performed_by': event.performed_by,
            'remarks': event.remarks,
        })
    
    # 2. AuditLog events involving this employee
    audit_events = AuditLog.query.filter(
        or_(
            AuditLog.employee_id == emp_id,
            AuditLog.employee_name == employee.employee_name
        )
    ).order_by(AuditLog.timestamp.desc()).all()
    
    for event in audit_events:
        asset = Asset.query.get(event.asset_id) if event.asset_id else None
        events.append({
            'type': 'audit',
            'action_type': event.action_type,
            'date': event.timestamp.isoformat() if event.timestamp else None,
            'timestamp': event.timestamp.isoformat() if event.timestamp else None,
            'asset_id': event.asset_id,
            'asset_name': event.asset_name or (asset.asset_name if asset else None),
            'asset_serial': event.asset_serial or (asset.serial_number if asset else None),
            'category': event.category or (asset.category if asset else None),
            'brand_name': asset.brand_name if asset else None,
            'model_name': asset.model_name if asset else None,
            'employee_name': event.employee_name,
            'employee_id': event.employee_id,
            'old_value': event.old_value,
            'new_value': event.new_value,
            'performed_by': event.performed_by,
            'remarks': event.remarks,
            'module': event.module,
        })
    
    # 3. Temporary assignments (both as recipient and as original owner)
    temp_assignments = TemporaryAssignment.query.filter_by(employee_id=emp_id).order_by(
        TemporaryAssignment.created_at.desc()
    ).all()
    
    for temp in temp_assignments:
        # Original asset event
        original_asset = Asset.query.get(temp.original_asset_id) if temp.original_asset_id else None
        events.append({
            'type': 'temp_assignment',
            'sub_type': 'original',
            'date': temp.start_date.isoformat() if temp.start_date else None,
            'timestamp': temp.created_at.isoformat() if temp.created_at else None,
            'asset_id': temp.original_asset_id,
            'asset_name': temp.original_asset_name or (original_asset.asset_name if original_asset else None),
            'asset_serial': temp.original_asset_serial or (original_asset.serial_number if original_asset else None),
            'category': original_asset.category if original_asset else None,
            'brand_name': original_asset.brand_name if original_asset else None,
            'model_name': original_asset.model_name if original_asset else None,
            'employee_name': temp.employee_name,
            'employee_id': temp.employee_id,
            'reason': temp.reason,
            'status': temp.status,
            'remarks': f"Original device sent for repair. Loaner: {temp.temp_asset_name}",
        })
        
        # Temporary replacement asset event
        temp_asset = Asset.query.get(temp.temp_asset_id) if temp.temp_asset_id else None
        events.append({
            'type': 'temp_assignment',
            'sub_type': 'temp',
            'date': temp.start_date.isoformat() if temp.start_date else None,
            'timestamp': temp.created_at.isoformat() if temp.created_at else None,
            'asset_id': temp.temp_asset_id,
            'asset_name': temp.temp_asset_name or (temp_asset.asset_name if temp_asset else None),
            'asset_serial': temp.temp_asset_serial or (temp_asset.serial_number if temp_asset else None),
            'category': temp_asset.category if temp_asset else None,
            'brand_name': temp_asset.brand_name if temp_asset else None,
            'model_name': temp_asset.model_name if temp_asset else None,
            'employee_name': temp.employee_name,
            'employee_id': temp.employee_id,
            'reason': temp.reason,
            'status': temp.status,
            'remarks': f"Temporary replacement while {temp.original_asset_name} is in repair",
        })
    
    # 4. Asset replacements (permanent swaps)
    replacements = AssetReplacement.query.filter_by(employee_id=emp_id).order_by(
        AssetReplacement.replacement_date.desc()
    ).all()
    
    for replacement in replacements:
        old_asset = Asset.query.get(replacement.old_asset_id) if replacement.old_asset_id else None
        new_asset = Asset.query.get(replacement.new_asset_id) if replacement.new_asset_id else None
        
        events.append({
            'type': 'replacement',
            'date': replacement.replacement_date.isoformat() if replacement.replacement_date else None,
            'timestamp': replacement.created_at.isoformat() if replacement.created_at else None,
            'old_asset_id': replacement.old_asset_id,
            'old_asset_name': replacement.old_asset_name or (old_asset.asset_name if old_asset else None),
            'old_asset_serial': replacement.old_asset_serial or (old_asset.serial_number if old_asset else None),
            'new_asset_id': replacement.new_asset_id,
            'asset_id': replacement.new_asset_id,  # For consistency
            'asset_name': replacement.new_asset_name or (new_asset.asset_name if new_asset else None),
            'asset_serial': replacement.new_asset_serial or (new_asset.serial_number if new_asset else None),
            'category': new_asset.category if new_asset else None,
            'brand_name': new_asset.brand_name if new_asset else None,
            'model_name': new_asset.model_name if new_asset else None,
            'employee_name': replacement.employee_name,
            'employee_id': replacement.employee_id,
            'reason': replacement.reason,
            'old_asset_condition': replacement.old_asset_condition,
            'performed_by': replacement.performed_by,
            'remarks': replacement.remarks,
        })
    
    # Sort all events by date (newest first)
    events.sort(key=lambda x: x.get('timestamp') or x.get('date') or '', reverse=True)
    
    # Calculate statistics
    total_assignments = len([e for e in events if e.get('type') in ['lifecycle', 'audit'] and 
                            e.get('event_type') == 'ASSIGNED' or e.get('action_type') == 'ASSET_ASSIGNED'])
    total_returns = len([e for e in events if e.get('type') in ['lifecycle', 'audit'] and 
                        e.get('event_type') == 'RETURNED' or e.get('action_type') == 'ASSET_RETURNED'])
    total_replacements = len([e for e in events if e.get('type') == 'replacement'])
    total_temp_assignments = len([e for e in events if e.get('type') == 'temp_assignment'])
    
    # Get unique assets this employee has used
    unique_asset_ids = set()
    for event in events:
        if event.get('asset_id'):
            unique_asset_ids.add(event.get('asset_id'))
    
    return jsonify({
        'employee': {
            'emp_id': employee.emp_id,
            'employee_name': employee.employee_name,
            'email': employee.email,
            'mobile_number': employee.mobile_number,
            'department': employee.department,
            'designation': employee.designation,
            'location': employee.location,
            'status': employee.status,
        },
        'current_assets': [{
            'id': a.id,
            'asset_name': a.asset_name,
            'category': a.category,
            'serial_number': a.serial_number,
            'model_name': a.model_name,
            'brand_name': a.brand_name,
            'status': a.status,
            'date': a.date.isoformat() if a.date else None,
        } for a in current_assets],
        'statistics': {
            'total_devices_used': len(unique_asset_ids),
            'current_assigned_devices': len(current_assets),
            'total_assignments': total_assignments,
            'total_returns': total_returns,
            'total_replacements': total_replacements,
            'total_temp_assignments': total_temp_assignments,
            'total_events': len(events),
        },
        'events': events,
    }), 200

@app.route('/api/employees/<emp_id>/exit', methods=['POST'])
@admin_required
def employee_exit(emp_id):
    """Process employee exit and asset recovery"""
    from models import Employee, Asset, AuditLog
    from datetime import datetime
    
    data = request.get_json()
    assets_recovery = data.get('assets', [])  # List of {asset_id, recovery_status, notes}
    exit_date_str = data.get('exit_date')
    exit_notes = data.get('exit_notes', '')
    
    # Parse exit date
    exit_date = None
    if exit_date_str:
        try:
            exit_date = datetime.strptime(exit_date_str, '%Y-%m-%d').date()
        except:
            exit_date = datetime.utcnow().date()
    else:
        exit_date = datetime.utcnow().date()
    
    # Get employee
    employee = Employee.query.filter_by(emp_id=emp_id).first()
    if not employee:
        return jsonify({'error': 'Employee not found'}), 404
    
    # Process each asset
    recovered_count = 0
    missing_count = 0
    damaged_count = 0
    
    for asset_recovery in assets_recovery:
        asset = Asset.query.get(asset_recovery['asset_id'])
        if not asset:
            continue
        
        recovery_status = asset_recovery['recovery_status']  # 'returned', 'missing', 'damaged'
        notes = asset_recovery.get('notes', '')
        
        # Update asset based on recovery status
        if recovery_status == 'returned':
            asset.status = 'Available'
            asset.emp_id = None
            asset.employee_name = None
            asset.employee_email = None
            asset.mobile_number = None
            recovered_count += 1
            
            # Log the return
            audit = AuditLog(
                asset_id=asset.id,
                action_type='ASSET_RETURNED',
                module='Asset',
                employee_name=employee.employee_name,
                old_value=f'Assigned to {employee.employee_name}',
                new_value='Available',
                remarks=f'Employee Exit: {notes}' if notes else 'Employee Exit - Asset Returned',
                performed_by='admin'
            )
            db.session.add(audit)
            
        elif recovery_status == 'missing':
            asset.status = 'Retired'  # Mark as retired since it's missing
            asset.emp_id = None
            asset.employee_name = None
            asset.employee_email = None
            asset.mobile_number = None
            missing_count += 1
            
            # Log as missing
            audit = AuditLog(
                asset_id=asset.id,
                action_type='ASSET_MISSING',
                module='Asset',
                employee_name=employee.employee_name,
                old_value=f'Assigned to {employee.employee_name}',
                new_value='Missing',
                remarks=f'Employee Exit: {notes}' if notes else 'Employee Exit - Asset Missing',
                performed_by='admin'
            )
            db.session.add(audit)
            
        elif recovery_status == 'damaged':
            asset.status = 'Maintenance'
            asset.emp_id = None
            asset.employee_name = None
            asset.employee_email = None
            asset.mobile_number = None
            damaged_count += 1
            
            # Log as damaged
            audit = AuditLog(
                asset_id=asset.id,
                action_type='ASSET_DAMAGED',
                module='Asset',
                employee_name=employee.employee_name,
                old_value=f'Assigned to {employee.employee_name}',
                new_value='Maintenance - Damaged',
                remarks=f'Employee Exit: {notes}' if notes else 'Employee Exit - Asset Damaged',
                performed_by='admin'
            )
            db.session.add(audit)
    
    # Mark employee as exited
    employee.status = 'Exited'
    employee.exit_date = exit_date
    
    # Create employee exit audit log
    exit_audit = AuditLog(
        asset_id=None,
        action_type='EMPLOYEE_EXIT',
        module='Employee',
        employee_name=employee.employee_name,
        old_value='Active',
        new_value='Exited',
        remarks=f'Exit Notes: {exit_notes}. Recovered: {recovered_count}, Missing: {missing_count}, Damaged: {damaged_count}',
        performed_by='admin'
    )
    db.session.add(exit_audit)
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Employee exit processed successfully',
        'summary': {
            'employee': employee.employee_name,
            'emp_id': emp_id,
            'recovered': recovered_count,
            'missing': missing_count,
            'damaged': damaged_count,
            'total_assets': len(assets_recovery)
        }
    }), 200

@app.route('/api/employees', methods=['POST'])
@admin_required
def create_or_update_employee():
    """Create or update employee"""
    from models import Employee
    
    data = request.get_json() or {}
    emp_id = data.get('emp_id', '').strip()
    
    if not emp_id:
        return jsonify({'error': 'emp_id is required'}), 400
    
    employee = Employee.query.filter_by(emp_id=emp_id).first()
    current_user = get_current_user()
    
    if employee:
        # Update existing
        employee.employee_name = data.get('employee_name', employee.employee_name)
        employee.email = data.get('email', employee.email)
        employee.mobile_number = data.get('mobile_number', employee.mobile_number)
        employee.location = data.get('location', employee.location)
        employee.department = data.get('department', employee.department)
        employee.designation = data.get('designation', employee.designation)
        employee.updated_at = datetime.utcnow()
        action = 'UPDATE'
    else:
        # Create new
        employee = Employee(
            emp_id=emp_id,
            employee_name=data.get('employee_name', ''),
            email=data.get('email', ''),
            mobile_number=data.get('mobile_number', ''),
            location=data.get('location', ''),
            department=data.get('department', ''),
            designation=data.get('designation', '')
        )
        db.session.add(employee)
        action = 'CREATE'
    
    log_activity(action, 'Employee', f'{action} employee: {employee.employee_name} [{emp_id}]', current_user)
    db.session.commit()
    
    return jsonify({'success': True, 'employee': {
        'emp_id': employee.emp_id,
        'employee_name': employee.employee_name,
        'email': employee.email,
        'mobile_number': employee.mobile_number,
        'location': employee.location,
    }}), 200 if action == 'UPDATE' else 201

# ══════════════════════════════════════════════════════════════════════════════
# WARRANTY ALERTS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/assets/warranty/expiring', methods=['GET'])
def warranty_expiring():
    days  = request.args.get('days', 90, type=int)
    today = date.today()
    soon  = today + timedelta(days=days)
    assets = Asset.query.filter(
        Asset.warranty_date != None,
        Asset.warranty_date <= soon,
        Asset.warranty_date >= today
    ).order_by(Asset.warranty_date).all()
    return jsonify({'assets': [a.to_dict() for a in assets]}), 200

# ══════════════════════════════════════════════════════════════════════════════
# REPORTS – CSV / Excel export
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/reports/export/csv', methods=['GET'])
def export_csv():
    assets = Asset.query.order_by(Asset.id).all()
    output = io.StringIO()
    writer = csv.writer(output)

    # Header – matches the 20 required columns
    writer.writerow([
        'Sl No', 'EMP ID', 'Employee Name', 'Mobile Number',
        'Asset Name', 'Category', 'Serial Number', 'Model Name',
        'OS', 'Version', 'RAM', 'Location',
        'Invoice Number', 'Invoice Date', 'Warranty Date',
        'Charger Serial Number', 'Old User', 'Date', 'Old Device',
        'Comments', 'Status'
    ])

    for a in assets:
        writer.writerow([
            a.id, a.emp_id, a.employee_name, a.mobile_number,
            a.asset_name, a.category, a.serial_number, a.model_name,
            a.os, a.version, a.ram, a.location,
            a.invoice_number, a.invoice_date, a.warranty_date,
            a.charger_serial, a.old_user, a.date, a.old_device,
            a.comments, a.status
        ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),  # utf-8-sig for Excel compatibility
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'IT_Assets_{date.today()}.csv'
    )

@app.route('/api/reports/export/excel', methods=['GET'])
def export_excel():
    """Export as Excel using openpyxl"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return export_csv()

    assets = Asset.query.order_by(Asset.id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'IT Assets'

    headers = [
        'Sl No', 'EMP ID', 'Employee Name', 'Mobile Number',
        'Asset Name', 'Category', 'Serial Number', 'Model Name',
        'OS', 'Version', 'RAM', 'Location',
        'Invoice Number', 'Invoice Date', 'Warranty Date',
        'Charger Serial Number', 'Old User', 'Date', 'Old Device',
        'Comments', 'Status'
    ]

    # Style header row
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 25

    # Data rows
    for row_idx, a in enumerate(assets, 2):
        row_data = [
            a.id, a.emp_id, a.employee_name, a.mobile_number,
            a.asset_name, a.category, a.serial_number, a.model_name,
            a.os, a.version, a.ram, a.location,
            a.invoice_number,
            a.invoice_date.strftime('%Y-%m-%d') if a.invoice_date else '',
            a.warranty_date.strftime('%Y-%m-%d') if a.warranty_date else '',
            a.charger_serial, a.old_user,
            a.date.strftime('%Y-%m-%d') if a.date else '',
            a.old_device, a.comments, a.status
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            # Alternate row color
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, len(assets) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'IT_Assets_{date.today()}.xlsx'
    )

@app.route('/api/reports/activity', methods=['GET'])
def activity_log():
    page     = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    total    = ActivityLog.query.count()
    logs     = ActivityLog.query.order_by(ActivityLog.timestamp.desc())\
                                .offset((page-1)*per_page).limit(per_page).all()
    return jsonify({
        'logs':  [l.to_dict() for l in logs],
        'total': total,
        'page':  page,
        'pages': (total + per_page - 1) // per_page,
    }), 200

# ── Audit Logs (Complete Activity History) ───────────────────────────────────
@app.route('/api/audit-logs', methods=['GET'])
def get_audit_logs():
    """Get audit logs with filters"""
    from models import AuditLog
    
    # Get filters (accept both date_from/date_to and start_date/end_date)
    action_type = request.args.get('action_type') or request.args.get('action')
    asset_id = request.args.get('asset_id')
    employee_id = request.args.get('employee_id')
    start_date = request.args.get('date_from') or request.args.get('start_date')
    end_date = request.args.get('date_to') or request.args.get('end_date')
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    q = AuditLog.query
    
    # Apply filters
    if action_type:
        q = q.filter_by(action_type=action_type)
    if asset_id:
        q = q.filter_by(asset_id=asset_id)
    if employee_id:
        q = q.filter_by(employee_id=employee_id)
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            q = q.filter(AuditLog.timestamp >= start)
        except:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            q = q.filter(AuditLog.timestamp <= end)
        except:
            pass
    if search:
        q = q.filter(or_(
            AuditLog.asset_name.ilike(f'%{search}%'),
            AuditLog.asset_serial.ilike(f'%{search}%'),
            AuditLog.employee_name.ilike(f'%{search}%'),
            AuditLog.employee_id.ilike(f'%{search}%'),
            AuditLog.remarks.ilike(f'%{search}%')
        ))
    
    total = q.count()
    logs = q.order_by(AuditLog.timestamp.desc()).offset((page-1)*per_page).limit(per_page).all()
    
    return jsonify({
        'logs': [{
            'id': log.id,
            'action_type': log.action_type,
            'module': log.module,
            'asset_id': log.asset_id,
            'asset_name': log.asset_name,
            'asset_serial': log.asset_serial,
            'category': log.category,
            'employee_id': log.employee_id,
            'employee_name': log.employee_name,
            'field_name': log.field_name,
            'old_value': log.old_value,
            'new_value': log.new_value,
            'performed_by': log.performed_by,
            'ip_address': log.ip_address,
            'remarks': log.remarks,
            'timestamp': log.timestamp.isoformat() if log.timestamp else None,
        } for log in logs],
        'total': total,
        'page': page,
        'pages': (total + per_page - 1) // per_page,
    }), 200

@app.route('/api/audit-logs/export', methods=['GET'])
def export_audit_logs():
    """Export audit logs to CSV"""
    from models import AuditLog
    
    # Get filters (same as get_audit_logs)
    action_type = request.args.get('action_type') or request.args.get('action')
    asset_id = request.args.get('asset_id')
    employee_id = request.args.get('employee_id')
    start_date = request.args.get('date_from') or request.args.get('start_date')
    end_date = request.args.get('date_to') or request.args.get('end_date')
    search = request.args.get('search', '').strip()
    
    q = AuditLog.query
    
    # Apply same filters
    if action_type:
        q = q.filter_by(action_type=action_type)
    if asset_id:
        q = q.filter_by(asset_id=asset_id)
    if employee_id:
        q = q.filter_by(employee_id=employee_id)
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            q = q.filter(AuditLog.timestamp >= start)
        except:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            q = q.filter(AuditLog.timestamp <= end)
        except:
            pass
    if search:
        q = q.filter(or_(
            AuditLog.asset_name.ilike(f'%{search}%'),
            AuditLog.asset_serial.ilike(f'%{search}%'),
            AuditLog.employee_name.ilike(f'%{search}%'),
            AuditLog.employee_id.ilike(f'%{search}%'),
            AuditLog.remarks.ilike(f'%{search}%')
        ))
    
    logs = q.order_by(AuditLog.timestamp.desc()).all()
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        'Timestamp', 'Action', 'Module', 'Asset Name', 'Serial Number',
        'Employee', 'Field', 'Old Value', 'New Value', 'Performed By',
        'IP Address', 'Remarks'
    ])
    
    # Data rows
    for log in logs:
        writer.writerow([
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else '',
            log.action_type,
            log.module,
            log.asset_name or '',
            log.asset_serial or '',
            log.employee_name or '',
            log.field_name or '',
            log.old_value or '',
            log.new_value or '',
            log.performed_by,
            log.ip_address or '',
            log.remarks or ''
        ])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'Activity_History_{date.today()}.csv'
    )

# ══════════════════════════════════════════════════════════════════════════════
# ASSET LIFECYCLE TIMELINE ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/lifecycle/asset/<int:asset_id>', methods=['GET'])
@non_viewer_required
def get_asset_lifecycle(asset_id):
    """Get complete lifecycle timeline for an asset"""
    from models import AssetLifecycle
    
    try:
        # Get all lifecycle events for this asset
        timeline = AssetLifecycle.query.filter_by(asset_id=asset_id).order_by(
            AssetLifecycle.event_date.desc()
        ).all()
        
        events = []
        for event in timeline:
            events.append({
                'id': event.id,
                'asset_id': event.asset_id,
                'event_type': event.event_type,
                'event_date': event.event_date.isoformat() if event.event_date else None,
                'from_employee_id': event.from_employee_id,
                'from_employee': event.from_employee,
                'to_employee_id': event.to_employee_id,
                'to_employee': event.to_employee,
                'from_status': event.from_status,
                'to_status': event.to_status,
                'reason': event.reason,
                'location': event.location,
                'performed_by': event.performed_by,
                'remarks': event.remarks,
                'created_at': event.created_at.isoformat() if event.created_at else None
            })
        
        return jsonify({
            'asset_id': asset_id,
            'events': events,
            'total': len(events)
        }), 200
        
    except Exception as e:
        logger.error(f"Error fetching asset lifecycle: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/lifecycle/holders/<int:asset_id>', methods=['GET'])
@non_viewer_required
def get_asset_holders(asset_id):
    """Get all employees who have held this asset"""
    from models import AssetLifecycle
    
    try:
        # Get unique employees who have held this asset
        holders_query = db.session.query(
            AssetLifecycle.to_employee_id,
            AssetLifecycle.to_employee,
            db.func.min(AssetLifecycle.event_date).label('first_assigned'),
            db.func.max(AssetLifecycle.event_date).label('last_event')
        ).filter(
            AssetLifecycle.asset_id == asset_id,
            AssetLifecycle.to_employee_id.isnot(None)
        ).group_by(
            AssetLifecycle.to_employee_id,
            AssetLifecycle.to_employee
        ).all()
        
        holders = []
        for holder in holders_query:
            holders.append({
                'employee_id': holder.to_employee_id,
                'employee_name': holder.to_employee,
                'first_assigned': holder.first_assigned.isoformat() if holder.first_assigned else None,
                'last_event': holder.last_event.isoformat() if holder.last_event else None
            })
        
        return jsonify({
            'asset_id': asset_id,
            'holders': holders,
            'total': len(holders)
        }), 200
        
    except Exception as e:
        logger.error(f"Error fetching asset holders: {e}")
        return jsonify({'error': str(e)}), 500

# ══════════════════════════════════════════════════════════════════════════════
# PDF GENERATION ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/assets/<int:asset_id>/assignment-form', methods=['GET'])
@token_required
def generate_assignment_form_pdf(asset_id):
    """Generate PDF assignment form for a single asset"""
    from models import Asset, Employee
    from services.pdf_generator import create_pdf_generator
    
    try:
        # Get asset details
        asset = Asset.query.get(asset_id)
        if not asset:
            return jsonify({'error': 'Asset not found'}), 404
        
        # Prepare asset data for PDF
        asset_data = {
            'asset_id': asset.id,
            'asset_name': asset.asset_name or 'N/A',
            'category': asset.category or 'N/A',
            'serial_number': asset.serial_number or 'N/A',
            'model': asset.model_name or 'N/A',
            'status': asset.status or 'N/A',
            'processor': asset.processor or 'N/A',
            'ram': asset.ram or 'N/A',
            'storage_capacity': asset.storage_capacity or 'N/A',
            'operating_system': asset.os or 'N/A',
            'invoice_number': asset.invoice_number or 'N/A',
            'invoice_date': asset.invoice_date.strftime('%d-%m-%Y') if asset.invoice_date else 'N/A',
            'warranty_date': asset.warranty_date.strftime('%d-%m-%Y') if asset.warranty_date else 'N/A',
            'charger_serial': asset.charger_serial or 'N/A',
            'assignment_date': asset.date.strftime('%d-%m-%Y') if asset.date else datetime.now().strftime('%d-%m-%Y'),
            'issued_by': 'Admin',
        }
        
        # Get employee details if assigned
        if asset.emp_id:
            employee = Employee.query.filter_by(emp_id=asset.emp_id).first()
            if employee:
                asset_data['employee_id'] = employee.emp_id
                asset_data['employee_name'] = employee.employee_name
                asset_data['department'] = employee.department or 'N/A'
                asset_data['mobile'] = employee.mobile_number or 'N/A'
                asset_data['email'] = employee.email or 'N/A'
                asset_data['location'] = employee.location or 'N/A'
            else:
                # Use data from asset if employee not in Employee table
                asset_data['employee_id'] = asset.emp_id
                asset_data['employee_name'] = asset.employee_name or 'N/A'
                asset_data['department'] = 'N/A'
                asset_data['mobile'] = asset.mobile_number or 'N/A'
                asset_data['email'] = asset.employee_email or 'N/A'
                asset_data['location'] = asset.location or 'N/A'
        else:
            asset_data['employee_id'] = 'N/A'
            asset_data['employee_name'] = 'Unassigned'
            asset_data['department'] = 'N/A'
            asset_data['mobile'] = 'N/A'
            asset_data['email'] = 'N/A'
            asset_data['location'] = asset.location or 'N/A'
        
        # Generate PDF
        pdf_generator = create_pdf_generator()
        pdf_bytes = pdf_generator.generate_assignment_form(asset_data)
        
        # Return PDF as response using send_file with BytesIO
        from flask import send_file
        pdf_buffer = io.BytesIO(pdf_bytes)
        pdf_buffer.seek(0)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'Assignment_Form_{asset_id}_{asset.asset_name or "Asset"}.pdf'.replace(' ', '_')
        )
        
    except Exception as e:
        logger.error(f"Error generating assignment form PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/assets/assignment-forms/bulk', methods=['POST'])
@token_required
def generate_bulk_assignment_forms():
    """Generate PDF assignment forms for multiple assets and return as ZIP"""
    from models import Asset, Employee
    from services.pdf_generator import create_pdf_generator
    
    try:
        data = request.json
        asset_ids = data.get('asset_ids', [])
        
        if not asset_ids:
            return jsonify({'error': 'No asset IDs provided'}), 400
        
        # Collect asset data
        assets_data = []
        for asset_id in asset_ids:
            asset = Asset.query.get(asset_id)
            if not asset:
                continue
            
            # Prepare asset data
            asset_data = {
                'asset_id': asset.id,
                'asset_name': asset.asset_name or 'N/A',
                'category': asset.category or 'N/A',
                'serial_number': asset.serial_number or 'N/A',
                'model': asset.model_name or 'N/A',
                'status': asset.status or 'N/A',
                'processor': asset.processor or 'N/A',
                'ram': asset.ram or 'N/A',
                'storage_capacity': asset.storage_capacity or 'N/A',
                'operating_system': asset.os or 'N/A',
                'invoice_number': asset.invoice_number or 'N/A',
                'invoice_date': asset.invoice_date.strftime('%d-%m-%Y') if asset.invoice_date else 'N/A',
                'warranty_date': asset.warranty_date.strftime('%d-%m-%Y') if asset.warranty_date else 'N/A',
                'charger_serial': asset.charger_serial or 'N/A',
                'assignment_date': asset.date.strftime('%d-%m-%Y') if asset.date else datetime.now().strftime('%d-%m-%Y'),
                'issued_by': 'Admin',
            }
            
            # Get employee details if assigned
            if asset.emp_id:
                employee = Employee.query.filter_by(emp_id=asset.emp_id).first()
                if employee:
                    asset_data['employee_id'] = employee.emp_id
                    asset_data['employee_name'] = employee.employee_name
                    asset_data['department'] = employee.department or 'N/A'
                    asset_data['mobile'] = employee.mobile_number or 'N/A'
                    asset_data['email'] = employee.email or 'N/A'
                    asset_data['location'] = employee.location or 'N/A'
                else:
                    # Use data from asset if employee not in Employee table
                    asset_data['employee_id'] = asset.emp_id
                    asset_data['employee_name'] = asset.employee_name or 'N/A'
                    asset_data['department'] = 'N/A'
                    asset_data['mobile'] = asset.mobile_number or 'N/A'
                    asset_data['email'] = asset.employee_email or 'N/A'
                    asset_data['location'] = asset.location or 'N/A'
            else:
                asset_data['employee_id'] = 'N/A'
                asset_data['employee_name'] = 'Unassigned'
                asset_data['department'] = 'N/A'
                asset_data['mobile'] = 'N/A'
                asset_data['email'] = 'N/A'
                asset_data['location'] = asset.location or 'N/A'
            
            assets_data.append(asset_data)
        
        if not assets_data:
            return jsonify({'error': 'No valid assets found'}), 404
        
        # Generate ZIP with all PDFs
        pdf_generator = create_pdf_generator()
        zip_bytes = pdf_generator.generate_bulk_assignment_forms(assets_data)
        
        # Return ZIP as response using send_file with BytesIO
        from flask import send_file
        zip_buffer = io.BytesIO(zip_bytes)
        zip_buffer.seek(0)
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'Assignment_Forms_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        )
        
    except Exception as e:
        logger.error(f"Error generating bulk assignment forms: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ══════════════════════════════════════════════════════════════════════════════
# GLOBAL SEARCH
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/search/global', methods=['GET'])
@admin_required
def global_search():
    """
    Global smart search across assets, employees, invoices, and inventory.
    Searches by: asset tag, serial, name, brand, model, category, employee name, 
    employee ID, email, invoice number, vendor, etc.
    """
    from models import Asset, Employee
    from sqlalchemy import or_, and_
    
    query = request.args.get('q', '').strip()
    filter_type = request.args.get('type', 'all').lower()  # all, assets, employees, inventory, invoices
    limit = request.args.get('limit', 10, type=int)
    
    if not query or len(query) < 2:
        return jsonify({
            'results': {
                'assets': [],
                'employees': [],
                'invoices': [],
                'inventory': []
            },
            'total': 0
        }), 200
    
    results = {
        'assets': [],
        'employees': [],
        'invoices': [],
        'inventory': []
    }
    
    search_pattern = f'%{query}%'
    
    # Search Assets
    if filter_type in ['all', 'assets', 'inventory']:
        try:
            asset_query = Asset.query.filter(
                or_(
                    Asset.serial_number.ilike(search_pattern),
                    Asset.asset_name.ilike(search_pattern),
                    Asset.brand_name.ilike(search_pattern),
                    Asset.model_name.ilike(search_pattern),
                    Asset.category.ilike(search_pattern),
                    Asset.ip_address.ilike(search_pattern),
                    Asset.location.ilike(search_pattern),
                    Asset.invoice_number.ilike(search_pattern),
                    Asset.purchase_vendor.ilike(search_pattern),
                )
            ).limit(limit).all()
            
            for asset in asset_query:
                result = {
                    'id': asset.id,
                    'type': 'asset',
                    'title': asset.asset_name,
                    'subtitle': f"Serial: {asset.serial_number}",
                    'category': asset.category,
                    'brand': asset.brand_name,
                    'model': asset.model_name,
                    'status': asset.status,
                    'location': asset.location,
                    'asset_tag': f"AST-{str(asset.id).zfill(5)}",
                    'url': f"/inventory/detail/{asset.id}"
                }
                
                # Add to inventory if available, otherwise to assets
                if asset.status == 'Available':
                    results['inventory'].append(result)
                else:
                    results['assets'].append(result)
        except Exception as e:
            print(f"Asset search error: {e}")
    
    # Search Employees
    if filter_type in ['all', 'employees']:
        try:
            # Get unique employees from assets
            employees_query = Asset.query.filter(
                and_(
                    Asset.emp_id.isnot(None),
                    or_(
                        Asset.emp_id.ilike(search_pattern),
                        Asset.employee_name.ilike(search_pattern),
                        Asset.employee_email.ilike(search_pattern),
                        Asset.mobile_number.ilike(search_pattern)
                    )
                )
            ).with_entities(
                Asset.emp_id,
                Asset.employee_name,
                Asset.employee_email,
                Asset.mobile_number
            ).distinct().limit(limit).all()
            
            seen_emp_ids = set()
            for emp in employees_query:
                if emp.emp_id and emp.emp_id not in seen_emp_ids:
                    seen_emp_ids.add(emp.emp_id)
                    results['employees'].append({
                        'type': 'employee',
                        'emp_id': emp.emp_id,
                        'title': emp.employee_name,
                        'subtitle': f"ID: {emp.emp_id}",
                        'email': emp.employee_email or '',
                        'mobile': emp.mobile_number or '',
                        'url': f"/employees/{emp.emp_id}/asset-history"
                    })
            
            # Also search in Employee table if it exists
            try:
                from models import Employee
                emp_table_query = Employee.query.filter(
                    or_(
                        Employee.emp_id.ilike(search_pattern),
                        Employee.employee_name.ilike(search_pattern),
                        Employee.email.ilike(search_pattern),
                        Employee.mobile_number.ilike(search_pattern),
                        Employee.department.ilike(search_pattern),
                        Employee.designation.ilike(search_pattern)
                    )
                ).limit(limit).all()
                
                for emp in emp_table_query:
                    if emp.emp_id not in seen_emp_ids:
                        seen_emp_ids.add(emp.emp_id)
                        results['employees'].append({
                            'type': 'employee',
                            'emp_id': emp.emp_id,
                            'title': emp.employee_name,
                            'subtitle': f"ID: {emp.emp_id}",
                            'email': emp.email or '',
                            'mobile': emp.mobile_number or '',
                            'department': emp.department or '',
                            'designation': emp.designation or '',
                            'url': f"/employees/{emp.emp_id}/asset-history"
                        })
            except:
                pass  # Employee table might not exist or be empty
                
        except Exception as e:
            print(f"Employee search error: {e}")
    
    # Search Invoices (via assets with invoice information)
    if filter_type in ['all', 'invoices']:
        try:
            invoice_query = Asset.query.filter(
                and_(
                    Asset.invoice_number.isnot(None),
                    or_(
                        Asset.invoice_number.ilike(search_pattern),
                        Asset.purchase_vendor.ilike(search_pattern)
                    )
                )
            ).with_entities(
                Asset.id,
                Asset.invoice_number,
                Asset.purchase_vendor,
                Asset.purchase_date,
                Asset.purchase_price,
                Asset.asset_name
            ).distinct().limit(limit).all()
            
            seen_invoices = set()
            for inv in invoice_query:
                if inv.invoice_number and inv.invoice_number not in seen_invoices:
                    seen_invoices.add(inv.invoice_number)
                    results['invoices'].append({
                        'type': 'invoice',
                        'invoice_number': inv.invoice_number,
                        'title': f"Invoice {inv.invoice_number}",
                        'subtitle': f"Vendor: {inv.purchase_vendor or 'N/A'}",
                        'vendor': inv.purchase_vendor,
                        'date': inv.purchase_date.isoformat() if inv.purchase_date else None,
                        'amount': inv.purchase_price,
                        'asset_id': inv.id,
                        'asset_name': inv.asset_name,
                        'url': f"/inventory/detail/{inv.id}"
                    })
        except Exception as e:
            print(f"Invoice search error: {e}")
    
    # Calculate total results
    total = (len(results['assets']) + len(results['employees']) + 
             len(results['invoices']) + len(results['inventory']))
    
    return jsonify({
        'results': results,
        'total': total,
        'query': query
    }), 200


# ── Health check & version info ──────────────────────────────────────────────
@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint"""
    db_status = 'healthy'
    try:
        # Check database connection with a simple query
        with app.app_context():
            result = db.session.execute(db.text('SELECT 1')).scalar()
            if result != 1:
                db_status = 'unhealthy'
    except Exception as e:
        logger.error(f"Database health check failed: {e}")
        db_status = 'unhealthy'
    
    return jsonify({
        'status': 'ok' if db_status == 'healthy' else 'degraded',
        'service': 'Tectoro Asset Management API',
        'version': '2.0.0',
        'database': db_status,
        'timestamp': datetime.utcnow().isoformat()
    }), 200

@app.route('/api/version', methods=['GET'])
def version():
    """API version information"""
    return jsonify({
        'version': '2.0.0',
        'api_version': 'v1',
        'name': 'Tectoro Asset Management API',
        'description': 'Complete IT Asset Management System',
        'security': 'JWT-based authentication',
        'features': [
            'Asset Management',
            'Employee Management',
            'Onboarding System',
            'Lifecycle Tracking',
            'Audit Logging',
            'Reports & Export',
            'Email Notifications'
        ]
    }), 200

# ══════════════════════════════════════════════════════════════════════════════
# ASSET REPLACEMENT ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/asset-replacements', methods=['GET'])
@token_required
def get_asset_replacements():
    """Get all asset replacements with optional filters"""
    from models import AssetReplacement
    
    status = request.args.get('status', '').strip()
    emp_id = request.args.get('employee_id', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    query = AssetReplacement.query
    
    if emp_id:
        query = query.filter_by(employee_id=emp_id)
    
    replacements = query.order_by(AssetReplacement.created_at.desc()).all()
    
    return jsonify({
        'success': True,
        'replacements': [r.to_dict() for r in replacements],
        'total': len(replacements)
    }), 200

@app.route('/api/asset-replacements', methods=['POST'])
@token_required
def create_asset_replacement():
    """Create a new asset replacement record"""
    from models import AssetReplacement
    
    data = request.get_json() or {}
    current_user = get_current_user()
    
    # Validate required fields
    required = ['employee_id', 'employee_name', 'old_asset_id', 'new_asset_id', 'reason']
    for field in required:
        if not data.get(field):
            return jsonify({'error': f'{field} is required'}), 400
    
    # Get asset details
    old_asset = Asset.query.get(data['old_asset_id'])
    new_asset = Asset.query.get(data['new_asset_id'])
    
    if not old_asset:
        return jsonify({'error': 'Old asset not found'}), 404
    if not new_asset:
        return jsonify({'error': 'New asset not found'}), 404
    
    # Create replacement record
    replacement = AssetReplacement(
        employee_id=data['employee_id'],
        employee_name=data['employee_name'],
        employee_email=data.get('employee_email'),
        old_asset_id=old_asset.id,
        old_asset_name=old_asset.asset_name,
        old_asset_serial=old_asset.serial_number,
        new_asset_id=new_asset.id,
        new_asset_name=new_asset.asset_name,
        new_asset_serial=new_asset.serial_number,
        replacement_date=parse_date(data.get('replacement_date')) or date.today(),
        reason=data['reason'],
        old_asset_condition=data.get('old_asset_condition', 'Good'),
        performed_by=current_user.get('username') if current_user else 'system',
        remarks=data.get('remarks', '')
    )
    db.session.add(replacement)
    
    # Update asset statuses
    old_asset.status = 'Retired'
    old_asset.emp_id = ''
    old_asset.employee_name = ''
    
    new_asset.status = 'Assigned'
    new_asset.emp_id = data['employee_id']
    new_asset.employee_name = data['employee_name']
    
    log_activity('CREATE', 'AssetReplacement', 
                f'Replaced asset for {data["employee_name"]}: {old_asset.asset_name} -> {new_asset.asset_name}',
                current_user.get('username') if current_user else 'system')
    db.session.commit()
    
    return jsonify({'success': True, 'replacement': replacement.to_dict()}), 201

@app.route('/api/asset-replacements/<int:replacement_id>', methods=['DELETE'])
@token_required
def delete_asset_replacement(replacement_id):
    """Delete an asset replacement record"""
    from models import AssetReplacement
    
    replacement = AssetReplacement.query.get_or_404(replacement_id)
    current_user = get_current_user()
    current_username = current_user.get('username') if current_user else 'system'
    
    employee_name = replacement.employee_name
    old_asset_name = replacement.old_asset_name
    new_asset_name = replacement.new_asset_name
    
    db.session.delete(replacement)
    log_activity('DELETE', 'AssetReplacement', 
                f'Deleted asset replacement for {employee_name}: {old_asset_name} -> {new_asset_name}',
                current_username)
    db.session.commit()
    
    return jsonify({'success': True}), 200

# ══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE EXIT ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/employee-exit', methods=['GET'])
@token_required
def get_employee_exits():
    """Get all employee exit records"""
    from models import EmployeeExit
    
    status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    query = EmployeeExit.query
    
    if status:
        query = query.filter_by(exit_status=status)
    
    exits = query.order_by(EmployeeExit.created_at.desc()).all()
    
    return jsonify({
        'success': True,
        'exits': [e.to_dict() for e in exits],
        'total': len(exits)
    }), 200

@app.route('/api/employee-exit/<int:exit_id>', methods=['GET'])
@token_required
def get_employee_exit_details(exit_id):
    """Get detailed employee exit information"""
    from models import EmployeeExit
    
    exit_record = EmployeeExit.query.get_or_404(exit_id)
    return jsonify({
        'success': True,
        'exit': exit_record.to_dict()
    }), 200

# ══════════════════════════════════════════════════════════════════════════════
# EMAIL CONFIGURATION ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/email-config', methods=['GET'])
@admin_required
def get_email_config():
    """Get email configuration (admin only)"""
    from models import EmailConfig
    
    config = EmailConfig.query.first()
    if not config:
        return jsonify({
            'success': True,
            'config': None,
            'message': 'No email configuration found'
        }), 200
    
    return jsonify({
        'success': True,
        'config': config.to_dict(include_password=False)
    }), 200

@app.route('/api/email-config', methods=['POST'])
@admin_required
def save_email_config():
    """Save email configuration (admin only)"""
    from models import EmailConfig
    
    data = request.get_json() or {}
    current_user = get_current_user()
    
    # Validate required fields
    required = ['sender_email', 'smtp_server', 'smtp_port', 'smtp_username', 'smtp_password']
    for field in required:
        if not data.get(field):
            return jsonify({'error': f'{field} is required'}), 400
    
    config = EmailConfig.query.first()
    if not config:
        config = EmailConfig()
        db.session.add(config)
    
    config.sender_email = data['sender_email']
    config.sender_name = data.get('sender_name', 'IT Asset Management')
    config.smtp_server = data['smtp_server']
    config.smtp_port = data['smtp_port']
    config.smtp_username = data['smtp_username']
    # Note: In production, encrypt this password
    config.smtp_password_enc = data['smtp_password']
    config.use_tls = data.get('use_tls', True)
    config.is_active = data.get('is_active', True)
    config.created_by = current_user.get('username') if current_user else 'admin'
    config.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Email configuration saved successfully',
        'config': config.to_dict(include_password=False)
    }), 200

@app.route('/api/email-config/test', methods=['POST'])
@admin_required
def test_email_config():
    """Test email configuration by sending a test email"""
    from models import EmailConfig
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    data = request.get_json() or {}
    test_recipient = data.get('test_recipient')
    
    if not test_recipient:
        return jsonify({'error': 'test_recipient email is required'}), 400
    
    config = EmailConfig.query.first()
    if not config:
        return jsonify({'error': 'No email configuration found'}), 404
    
    try:
        # Create test email
        msg = MIMEMultipart()
        msg['From'] = f"{config.sender_name} <{config.sender_email}>"
        msg['To'] = test_recipient
        msg['Subject'] = "Test Email from Asset Management System"
        
        body = """
        <html>
        <body>
        <h2>Test Email Successful!</h2>
        <p>This is a test email from your IT Asset Management System.</p>
        <p>If you received this email, your email configuration is working correctly.</p>
        </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))
        
        # Send email
        server = smtplib.SMTP(config.smtp_server, config.smtp_port)
        if config.use_tls:
            server.starttls()
        server.login(config.smtp_username, config.smtp_password_enc)
        server.send_message(msg)
        server.quit()
        
        # Update test status
        config.last_tested_at = datetime.utcnow()
        config.last_test_status = 'success'
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Test email sent successfully to {test_recipient}'
        }), 200
        
    except Exception as e:
        logger.error(f"Email test failed: {e}")
        
        # Update test status
        if config:
            config.last_tested_at = datetime.utcnow()
            config.last_test_status = 'failed'
            db.session.commit()
        
        return jsonify({
            'success': False,
            'error': f'Email test failed: {str(e)}'
        }), 500

# ══════════════════════════════════════════════════════════════════════════════
# LEGACY ENDPOINT ALIASES (for backward compatibility)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/activity-log', methods=['GET'])
@token_required
def get_activity_log_legacy():
    """Legacy alias for /api/audit-logs"""
    return get_audit_logs()

@app.route('/api/audit-log', methods=['GET'])
@token_required
def get_audit_log_legacy():
    """Legacy alias for /api/audit-logs"""
    return get_audit_logs()

# ── Serve React Frontend via 404 handler ────────────────────────────────────
@app.errorhandler(404)
def page_not_found(e):
    """Serve React app for 404 errors (SPA routing)"""
    path = request.path
    
    # If it's an API route, return actual 404
    if path.startswith('/api/'):
        logger.debug(f"API route not found: {path}")
        return jsonify({'error': 'Endpoint not found', 'path': path}), 404
    
    # Otherwise serve React app with no-cache headers for SPA routing
    logger.debug(f"Serving React SPA for path: {path}")
    try:
        response = send_from_directory(app.static_folder, 'index.html')
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        logger.error(f"Error serving index.html: {e}")
        return jsonify({'error': 'Application not found'}), 404

# ══════════════════════════════════════════════════════════════════════════════
# DATABASE INITIALIZATION
# ══════════════════════════════════════════════════════════════════════════════
# The office database (APP_ENV=office) is NEVER seeded — it only ever holds
# real office data, created and edited through the application itself.
# The public demo database (APP_ENV=render) is seeded with clearly-synthetic
# demo data the first time it's empty, so the public deployment always has
# something to show without ever touching real office data.
# ══════════════════════════════════════════════════════════════════════════════

with app.app_context():
    db.create_all()  # Create tables if they don't exist
    if is_render_env(APP_ENV):
        from demo_seed import seed_demo_data
        if seed_demo_data(db):
            logger.info("Demo database was empty — seeded with synthetic demo data")
        else:
            logger.info("Demo database already has data — skipped seeding")
    else:
        logger.info("Database tables initialized (office database — seeding never runs)")

# ONBOARDING API ROUTES — append this block to api_server.py
# Paste this near your other @app.route('/api/...') definitions, AFTER the
# imports at the top already bring in: app, db, jsonify, request, Asset,
# Employee. Add the import line below if Onboarding/OnboardingAssetAssignment
# aren't already imported.
# ─────────────────────────────────────────────────────────────────────────────

# Add this import near your other model imports at the top of api_server.py:
# from models import Onboarding, OnboardingAssetAssignment

from datetime import datetime


def _validate_onboarding_payload(data, is_update=False, current_id=None):
    """Returns (errors_dict, cleaned_data). errors_dict is empty if valid."""
    errors = {}

    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip()
    phone = (data.get('phone_number') or '').strip()
    designation = (data.get('designation') or '').strip()
    team = (data.get('team') or '').strip()

    if not name:
        errors['name'] = 'Name is required'
    if not email:
        errors['email'] = 'Email is required'
    if not phone:
        errors['phone_number'] = 'Phone number is required'
    if not designation:
        errors['designation'] = 'Designation is required'
    if not team:
        errors['team'] = 'Team is required'

    # Uniqueness check on email (excluding current record on update)
    if email and not errors.get('email'):
        existing = Onboarding.query.filter_by(email=email).first()
        if existing and (not is_update or existing.id != current_id):
            errors['email'] = 'An onboarding record with this email already exists'

    application_access = data.get('application_access', [])
    if isinstance(application_access, list):
        application_access_str = ','.join(application_access)
    else:
        application_access_str = str(application_access or '')

    cleaned = {
        'name': name,
        'email': email,
        'phone_number': phone,
        'designation': designation,
        'team': team,
        'application_access': application_access_str,
    }
    return errors, cleaned


# ── CREATE ──────────────────────────────────────────────────────────────────
@app.route('/api/onboarding', methods=['POST'])
@admin_required
def create_onboarding():
    data = request.get_json() or {}
    errors, cleaned = _validate_onboarding_payload(data)
    if errors:
        return jsonify({'success': False, 'errors': errors}), 400

    record = Onboarding(
        name=cleaned['name'],
        email=cleaned['email'],
        phone_number=cleaned['phone_number'],
        designation=cleaned['designation'],
        team=cleaned['team'],
        application_access=cleaned['application_access'],
        status=data.get('status', 'Pending'),
    )
    db.session.add(record)
    db.session.commit()

    # Optional: assign assets at creation time
    asset_ids = data.get('asset_ids', [])
    for asset_id in asset_ids:
        asset = Asset.query.get(asset_id)
        if asset:
            assignment = OnboardingAssetAssignment(
                onboarding_id=record.id,
                asset_id=asset.id,
                asset_name=asset.asset_name,
                asset_serial=asset.serial_number,
                asset_category=asset.category,
            )
            db.session.add(assignment)
    if asset_ids:
        db.session.commit()

    return jsonify({'success': True, 'onboarding': record.to_dict()}), 201


# ── LIST (with search, filter, sort, pagination) ─────────────────────────────
@app.route('/api/onboarding', methods=['GET'])
@admin_required
def list_onboarding():
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip()
    team = request.args.get('team', '').strip()
    sort = request.args.get('sort', 'created_desc')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    query = Onboarding.query

    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Onboarding.name.ilike(like),
                Onboarding.email.ilike(like),
                Onboarding.phone_number.ilike(like),
            )
        )
    if status:
        query = query.filter(Onboarding.status == status)
    if team:
        query = query.filter(Onboarding.team == team)

    sort_map = {
        'created_desc': Onboarding.created_at.desc(),
        'created_asc': Onboarding.created_at.asc(),
        'name_asc': Onboarding.name.asc(),
        'name_desc': Onboarding.name.desc(),
    }
    query = query.order_by(sort_map.get(sort, Onboarding.created_at.desc()))

    paginated = query.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        'success': True,
        'records': [r.to_dict() for r in paginated.items],
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': page,
    })


# ── GET single record ────────────────────────────────────────────────────────
@app.route('/api/onboarding/<int:onboarding_id>', methods=['GET'])
@admin_required
def get_onboarding(onboarding_id):
    record = Onboarding.query.get(onboarding_id)
    if not record:
        return jsonify({'success': False, 'error': 'Onboarding record not found'}), 404
    return jsonify({'success': True, 'onboarding': record.to_dict()})


# ── UPDATE ────────────────────────────────────────────────────────────────────
@app.route('/api/onboarding/<int:onboarding_id>', methods=['PUT'])
@admin_required
def update_onboarding(onboarding_id):
    record = Onboarding.query.get(onboarding_id)
    if not record:
        return jsonify({'success': False, 'error': 'Onboarding record not found'}), 404

    data = request.get_json() or {}
    errors, cleaned = _validate_onboarding_payload(data, is_update=True, current_id=onboarding_id)
    if errors:
        return jsonify({'success': False, 'errors': errors}), 400

    record.name = cleaned['name']
    record.email = cleaned['email']
    record.phone_number = cleaned['phone_number']
    record.designation = cleaned['designation']
    record.team = cleaned['team']
    record.application_access = cleaned['application_access']
    if 'status' in data:
        record.status = data['status']
    record.updated_at = datetime.utcnow()

    # Replace asset assignments if asset_ids provided
    if 'asset_ids' in data:
        OnboardingAssetAssignment.query.filter_by(onboarding_id=record.id).delete()
        for asset_id in data['asset_ids']:
            asset = Asset.query.get(asset_id)
            if asset:
                db.session.add(OnboardingAssetAssignment(
                    onboarding_id=record.id,
                    asset_id=asset.id,
                    asset_name=asset.asset_name,
                    asset_serial=asset.serial_number,
                    asset_category=asset.category,
                ))

    db.session.commit()
    return jsonify({'success': True, 'onboarding': record.to_dict()})


# ── DELETE ────────────────────────────────────────────────────────────────────
@app.route('/api/onboarding/<int:onboarding_id>', methods=['DELETE'])
@admin_required
def delete_onboarding(onboarding_id):
    record = Onboarding.query.get(onboarding_id)
    if not record:
        return jsonify({'success': False, 'error': 'Onboarding record not found'}), 404

    if record.status == 'Converted':
        return jsonify({
            'success': False,
            'error': 'Cannot delete a converted onboarding record. It is linked to an active employee.'
        }), 403

    db.session.delete(record)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Onboarding record deleted'})


# ── CONVERT TO ACTIVE EMPLOYEE ───────────────────────────────────────────────
# This satisfies: "Ensure a newly onboarded employee can be converted into an
# active employee without duplicate data entry."
@app.route('/api/onboarding/<int:onboarding_id>/convert', methods=['POST'])
@admin_required
def convert_onboarding_to_employee(onboarding_id):
    record = Onboarding.query.get(onboarding_id)
    if not record:
        return jsonify({'success': False, 'error': 'Onboarding record not found'}), 404

    if record.status == 'Converted':
        return jsonify({'success': False, 'error': 'This record has already been converted'}), 400

    data = request.get_json() or {}
    emp_id = (data.get('emp_id') or '').strip()
    if not emp_id:
        return jsonify({'success': False, 'error': 'emp_id is required to create the employee record'}), 400

    if Employee.query.filter_by(emp_id=emp_id).first():
        return jsonify({'success': False, 'error': f'Employee ID {emp_id} already exists'}), 400

    # Create the real Employee record — no re-typing of name/email/phone/etc.
    employee = Employee(
        emp_id=emp_id,
        employee_name=record.name,
        email=record.email,
        mobile_number=record.phone_number,
        designation=record.designation,
        department=record.team,
        status='Active',
        is_active=True,
    )
    db.session.add(employee)

    # Re-point asset assignments from onboarding -> the real employee
    for assignment in record.asset_assignments:
        asset = Asset.query.get(assignment.asset_id)
        if asset:
            asset.emp_id = emp_id
            asset.employee_name = record.name
            asset.employee_email = record.email
            asset.status = 'Assigned'

    record.status = 'Converted'
    record.converted_emp_id = emp_id
    record.converted_at = datetime.utcnow()

    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'{record.name} converted to active employee {emp_id}',
        'employee': employee.to_dict(),
        'onboarding': record.to_dict(),
    })


# ── ASSET PICKER HELPER — search available assets to assign during onboarding
@app.route('/api/onboarding/available-assets', methods=['GET'])
@admin_required
def onboarding_available_assets():
    search = request.args.get('search', '').strip()
    category = request.args.get('category', '').strip()

    query = Asset.query.filter(Asset.status == 'Available')
    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Asset.asset_name.ilike(like),
                Asset.serial_number.ilike(like),
            )
        )
    if category:
        query = query.filter(Asset.category == category)

    assets = query.order_by(Asset.asset_name).limit(50).all()
    return jsonify({
        'success': True,
        'assets': [{
            'id': a.id,
            'asset_name': a.asset_name,
            'serial_number': a.serial_number,
            'category': a.category,
            'status': a.status,
        } for a in assets]
    })

# ─────────────────────────────────────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════════════════
# CORPORATE SIM ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/corporate-sims', methods=['GET'])
def get_corporate_sims():
    """Get all Corporate SIMs with pagination, search, and filters"""
    from models import CorporateSIM
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    # Search query
    search = request.args.get('search', '').strip()
    
    # Filters
    status = request.args.get('status', '').strip()
    carrier = request.args.get('carrier', '').strip()
    assigned_to = request.args.get('assigned_to', '').strip()
    
    # Build query
    query = CorporateSIM.query
    
    # Apply search
    if search:
        query = query.filter(or_(
            CorporateSIM.iccid.ilike(f'%{search}%'),
            CorporateSIM.mobile_number.ilike(f'%{search}%'),
            CorporateSIM.assigned_employee_name.ilike(f'%{search}%'),
            CorporateSIM.corporate_account.ilike(f'%{search}%')
        ))
    
    # Apply filters
    if status:
        query = query.filter_by(status=status)
    if carrier:
        query = query.filter_by(carrier=carrier)
    if assigned_to:
        query = query.filter(or_(
            CorporateSIM.assigned_employee_id.ilike(f'%{assigned_to}%'),
            CorporateSIM.assigned_employee_name.ilike(f'%{assigned_to}%')
        ))
    
    # Get total count
    total = query.count()
    
    # Apply pagination and get results
    sims = query.order_by(CorporateSIM.created_at.desc())\
                .offset((page - 1) * per_page)\
                .limit(per_page)\
                .all()
    
    return jsonify({
        'sims': [sim.to_dict() for sim in sims],
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': (total + per_page - 1) // per_page
    })

@app.route('/api/corporate-sims/<int:sim_id>', methods=['GET'])
def get_corporate_sim(sim_id):
    """Get a single Corporate SIM by ID"""
    from models import CorporateSIM
    
    sim = CorporateSIM.query.get(sim_id)
    if not sim:
        return jsonify({'error': 'SIM not found'}), 404
    
    return jsonify({'sim': sim.to_dict()})

@app.route('/api/corporate-sims', methods=['POST'])
@admin_required
def create_corporate_sim():
    """Create a new Corporate SIM"""
    from models import CorporateSIM
    
    data = request.get_json() or {}
    
    # Validate required fields
    if not data.get('iccid'):
        return jsonify({'error': 'ICCID is required'}), 400
    if not data.get('carrier'):
        return jsonify({'error': 'Carrier is required'}), 400
    
    # Validate ICCID format (19-20 digits)
    iccid = data['iccid'].strip()
    if not iccid.isdigit() or len(iccid) not in [19, 20]:
        return jsonify({'error': 'ICCID must be 19-20 digits'}), 400
    
    # Check for duplicate ICCID
    if CorporateSIM.query.filter_by(iccid=iccid).first():
        return jsonify({'error': f'SIM with ICCID {iccid} already exists'}), 400
    
    # Validate mobile number if provided
    mobile = data.get('mobile_number', '').strip()
    if mobile:
        if not mobile.isdigit() or len(mobile) != 10:
            return jsonify({'error': 'Mobile number must be 10 digits'}), 400
        # Check for duplicate mobile number
        if CorporateSIM.query.filter_by(mobile_number=mobile).first():
            return jsonify({'error': f'SIM with mobile number {mobile} already exists'}), 400
    
    # Create new SIM
    current_user = get_current_user()
    sim = CorporateSIM(
        iccid=iccid,
        mobile_number=mobile if mobile else None,
        carrier=data.get('carrier'),
        plan_type=data.get('plan_type'),
        monthly_cost=data.get('monthly_cost'),
        data_limit_gb=data.get('data_limit_gb'),
        corporate_account=data.get('corporate_account'),
        account_manager=data.get('account_manager'),
        status=data.get('status', 'Available'),
        purchase_date=datetime.strptime(data['purchase_date'], '%Y-%m-%d').date() if data.get('purchase_date') else None,
        activation_date=datetime.strptime(data['activation_date'], '%Y-%m-%d').date() if data.get('activation_date') else None,
        vendor=data.get('vendor'),
        sim_type=data.get('sim_type'),
        puk_code=data.get('puk_code'),
        remarks=data.get('remarks'),
        created_by=current_user.username if current_user else 'system'
    )
    
    db.session.add(sim)
    db.session.commit()
    
    log_activity('CREATE', 'CorporateSIM', f'Created SIM: {iccid} - {data.get("carrier")}', current_user)
    
    return jsonify({'success': True, 'sim': sim.to_dict()}), 201

@app.route('/api/corporate-sims/<int:sim_id>', methods=['PUT'])
@admin_required
def update_corporate_sim(sim_id):
    """Update a Corporate SIM"""
    from models import CorporateSIM
    
    sim = CorporateSIM.query.get(sim_id)
    if not sim:
        return jsonify({'error': 'SIM not found'}), 404
    
    data = request.get_json() or {}
    current_user = get_current_user()
    
    # Update fields
    if 'mobile_number' in data:
        mobile = data['mobile_number'].strip()
        if mobile:
            if not mobile.isdigit() or len(mobile) != 10:
                return jsonify({'error': 'Mobile number must be 10 digits'}), 400
            # Check for duplicate (excluding current SIM)
            existing = CorporateSIM.query.filter_by(mobile_number=mobile).first()
            if existing and existing.id != sim_id:
                return jsonify({'error': f'Mobile number {mobile} is already in use'}), 400
            sim.mobile_number = mobile
        else:
            sim.mobile_number = None
    
    if 'carrier' in data:
        sim.carrier = data['carrier']
    if 'plan_type' in data:
        sim.plan_type = data['plan_type']
    if 'monthly_cost' in data:
        sim.monthly_cost = data['monthly_cost']
    if 'data_limit_gb' in data:
        sim.data_limit_gb = data['data_limit_gb']
    if 'corporate_account' in data:
        sim.corporate_account = data['corporate_account']
    if 'account_manager' in data:
        sim.account_manager = data['account_manager']
    if 'status' in data:
        sim.status = data['status']
    if 'purchase_date' in data:
        sim.purchase_date = datetime.strptime(data['purchase_date'], '%Y-%m-%d').date() if data['purchase_date'] else None
    if 'activation_date' in data:
        sim.activation_date = datetime.strptime(data['activation_date'], '%Y-%m-%d').date() if data['activation_date'] else None
    if 'vendor' in data:
        sim.vendor = data['vendor']
    if 'sim_type' in data:
        sim.sim_type = data['sim_type']
    if 'puk_code' in data:
        sim.puk_code = data['puk_code']
    if 'remarks' in data:
        sim.remarks = data['remarks']
    
    sim.updated_by = current_user.username if current_user else 'system'
    sim.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    log_activity('UPDATE', 'CorporateSIM', f'Updated SIM: {sim.iccid}', current_user)
    
    return jsonify({'success': True, 'sim': sim.to_dict()})

@app.route('/api/corporate-sims/<int:sim_id>', methods=['DELETE'])
@admin_required
def delete_corporate_sim(sim_id):
    """Delete a Corporate SIM (admin only)"""
    from models import CorporateSIM
    
    sim = CorporateSIM.query.get(sim_id)
    if not sim:
        return jsonify({'error': 'SIM not found'}), 404
    
    # Prevent deletion if assigned
    if sim.status == 'Assigned':
        return jsonify({'error': 'Cannot delete assigned SIM. Please return it first.'}), 400
    
    iccid = sim.iccid
    current_user = get_current_user()
    
    db.session.delete(sim)
    db.session.commit()
    
    log_activity('DELETE', 'CorporateSIM', f'Deleted SIM: {iccid}', current_user)
    
    return jsonify({'success': True, 'message': 'SIM deleted successfully'})

@app.route('/api/corporate-sims/<int:sim_id>/assign', methods=['POST'])
@admin_required
def assign_corporate_sim(sim_id):
    """Assign a Corporate SIM to an employee"""
    from models import CorporateSIM, Employee
    
    sim = CorporateSIM.query.get(sim_id)
    if not sim:
        return jsonify({'error': 'SIM not found'}), 404
    
    if sim.status not in ['Available', 'Returned']:
        return jsonify({'error': f'SIM is not available for assignment (current status: {sim.status})'}), 400
    
    data = request.get_json() or {}
    employee_id = data.get('employee_id', '').strip()
    
    if not employee_id:
        return jsonify({'error': 'Employee ID is required'}), 400
    
    # Get employee details
    employee = Employee.query.filter_by(emp_id=employee_id).first()
    if not employee:
        return jsonify({'error': f'Employee {employee_id} not found'}), 404
    
    # Assign SIM
    current_user = get_current_user()
    sim.assigned_employee_id = employee.emp_id
    sim.assigned_employee_name = employee.employee_name
    sim.assigned_employee_email = employee.email
    sim.assignment_date = date.today()
    sim.return_date = None
    sim.status = 'Assigned'
    sim.updated_by = current_user.username if current_user else 'system'
    sim.updated_at = datetime.utcnow()
    
    if data.get('remarks'):
        sim.remarks = (sim.remarks or '') + f"\n[{date.today()}] Assigned to {employee.employee_name}: {data['remarks']}"
    
    db.session.commit()
    
    log_activity('ASSIGN', 'CorporateSIM', f'Assigned SIM {sim.iccid} to {employee.employee_name} [{employee_id}]', current_user)
    
    return jsonify({'success': True, 'sim': sim.to_dict()})

@app.route('/api/corporate-sims/<int:sim_id>/return', methods=['POST'])
@admin_required
def return_corporate_sim(sim_id):
    """Return a Corporate SIM from an employee"""
    from models import CorporateSIM
    
    sim = CorporateSIM.query.get(sim_id)
    if not sim:
        return jsonify({'error': 'SIM not found'}), 404
    
    if sim.status != 'Assigned':
        return jsonify({'error': f'SIM is not assigned (current status: {sim.status})'}), 400
    
    data = request.get_json() or {}
    current_user = get_current_user()
    
    # Record return
    old_employee = sim.assigned_employee_name
    sim.return_date = date.today()
    sim.status = data.get('new_status', 'Available')  # Can be Available, Damaged, Lost, etc.
    
    # Clear assignment if returning to available
    if sim.status == 'Available':
        sim.assigned_employee_id = None
        sim.assigned_employee_name = None
        sim.assigned_employee_email = None
    
    if data.get('remarks'):
        sim.remarks = (sim.remarks or '') + f"\n[{date.today()}] Returned from {old_employee}: {data['remarks']}"
    
    sim.updated_by = current_user.username if current_user else 'system'
    sim.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    log_activity('RETURN', 'CorporateSIM', f'Returned SIM {sim.iccid} from {old_employee}', current_user)
    
    return jsonify({'success': True, 'sim': sim.to_dict()})

@app.route('/api/corporate-sims/stats', methods=['GET'])
def get_corporate_sim_stats():
    """Get Corporate SIM statistics for dashboard"""
    from models import CorporateSIM
    
    total = CorporateSIM.query.count()
    available = CorporateSIM.query.filter_by(status='Available').count()
    assigned = CorporateSIM.query.filter_by(status='Assigned').count()
    suspended = CorporateSIM.query.filter_by(status='Suspended').count()
    lost = CorporateSIM.query.filter_by(status='Lost').count()
    damaged = CorporateSIM.query.filter_by(status='Damaged').count()
    
    # Carrier breakdown
    carrier_stats = db.session.query(
        CorporateSIM.carrier,
        func.count(CorporateSIM.id).label('count')
    ).group_by(CorporateSIM.carrier).all()
    
    return jsonify({
        'total': total,
        'available': available,
        'assigned': assigned,
        'suspended': suspended,
        'lost': lost,
        'damaged': damaged,
        'carriers': [{'name': c[0], 'count': c[1]} for c in carrier_stats]
    })
if __name__ == '__main__':
    print("=" * 60)
    print("🚀  IT Asset Management API")
    print("=" * 60)
    print("✅  API:    http://0.0.0.0:3000")
    print("✅  Health: http://localhost:3000/api/health")
    print("⚛️   React:  Served from /frontend/build")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=3000)
